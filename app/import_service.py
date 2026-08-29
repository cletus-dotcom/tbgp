from datetime import datetime
from io import BytesIO
from math import isnan
from pathlib import Path

from app import db
from app.config import (
    MEMBER_LIFETIME_EARNINGS_CAP,
    MEMBER_SEPARATION_TYPES,
    MEMBER_STATUSES,
    MEMBERS_SHEET,
    MEMBERS_XLSX,
    is_admin_role,
)
from app.models import Member

REQUIRED_COLUMNS = {
    "member_id",
    "batch",
    "referrer_id",
    "date_joined",
    "last_name",
    "first_name",
    "middle_name",
    "suffix",
    "address",
    "phone",
    "email",
    "birth_date",
    "gender",
    "civil_status",
    "highest_education",
    "occupation_income_source",
    "monthly_income",
    "number_of_dependents",
    "beneficiary_name",
    "beneficiary_address",
    "beneficiary_phone",
    "status",
    "termination_date",
    "termination_type",
}

OPTIONAL_COLUMNS = {
    "membership_type",
    "lifetime_cap_enabled",
    "lifetime_cap_amount",
    "gcash_number",
    "bank_account_number",
    "bank_name",
    "age",
    "id_picture_location",
    "beneficiary_relationship",
}


class SheetRows:
    """Lightweight tabular sheet data without building a pandas DataFrame."""

    def __init__(self, rows):
        self._rows = rows

    @property
    def columns(self):
        if not self._rows:
            return []
        return list(self._rows[0].keys())

    @property
    def empty(self):
        return not self._rows

    def to_dict(self, orient="records"):
        if orient != "records":
            raise ValueError("Only orient='records' is supported.")
        return list(self._rows)


def _is_blank(value):
    if value is None:
        return True
    if isinstance(value, float) and isnan(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _parse_date(value):
    if _is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "date") and not isinstance(value, str):
        try:
            return value.date()
        except (AttributeError, TypeError, ValueError):
            pass

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _clean_str(value):
    if _is_blank(value):
        return None
    text = str(value).strip()
    return text or None


def _parse_int(value):
    if _is_blank(value):
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def _parse_dependents(value):
    parsed = _parse_int(value)
    if parsed is None:
        return None
    if parsed < 0 or parsed > 50:
        return None
    return parsed


def _parse_age(value):
    if _is_blank(value):
        return None
    parsed = _parse_int(value)
    if parsed is not None:
        return parsed if 0 <= parsed <= 130 else None
    text = str(value).strip()
    digits = "".join(ch for ch in text if ch.isdigit())
    if not digits:
        return None
    try:
        age = int(digits)
    except ValueError:
        return None
    return age if 0 <= age <= 130 else None


def _parse_bool(value, default=True):
    if _is_blank(value):
        return default
    text = str(value).strip().lower()
    if not text:
        return default
    return text in ("1", "true", "yes", "y", "on")


def _parse_decimal(value, default):
    if _is_blank(value):
        return default
    text = str(value).strip().replace(",", "")
    if not text:
        return default
    return float(text)


def _read_excel_openpyxl(source, sheet_name=None):
    import openpyxl

    if isinstance(source, (str, Path)):
        workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    else:
        content = source.read()
        if hasattr(source, "seek"):
            source.seek(0)
        workbook = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)

    if sheet_name and sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f'Sheet "{sheet_name}" was not found.')

    target_sheet = sheet_name if sheet_name in workbook.sheetnames else workbook.sheetnames[0]
    worksheet = workbook[target_sheet]
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return SheetRows([])

    headers = [str(col).strip() if col is not None else "" for col in rows[0]]
    records = []
    for row in rows[1:]:
        if row is None or all(_is_blank(cell) for cell in row):
            continue
        record = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            record[header] = row[index] if index < len(row) else None
        if _is_blank(record.get("member_id")) and _is_blank(record.get("contractor_id")) and _is_blank(record.get("supplier_id")):
            first_value = next((record[key] for key in record if not _is_blank(record.get(key))), None)
            if first_value is None:
                continue
        records.append(record)
    return SheetRows(records)


def _read_excel(source, sheet_name=None):
    """Read spreadsheet rows via openpyxl (avoids pandas Excel engine crashes)."""
    return _read_excel_openpyxl(source, sheet_name=sheet_name)


def load_members_dataframe(source):
    try:
        sheet = _read_excel(source, sheet_name=MEMBERS_SHEET)
    except ValueError:
        sheet = _read_excel(source)

    missing = REQUIRED_COLUMNS - set(sheet.columns)
    if missing:
        raise ValueError(f"Missing columns in spreadsheet: {', '.join(sorted(missing))}")

    if sheet.empty:
        raise ValueError("The spreadsheet has no data rows.")

    return sheet


def _clear_members_for_replace():
    """Remove member rows and dependent records that block a full replace import."""
    from app.models import (
        AdSplitMember,
        Contractor,
        MarketplaceLead,
        MemberLedger,
        OmpdFundEntry,
        PayoutNotification,
        PayoutRequest,
        ProductCommission,
        ProductCommissionAdAllocation,
        ProductCommissionShare,
        ProjectBilling,
        ProjectCommission,
        SharingBatch,
        SharingEntry,
        Supplier,
        User,
    )

    User.query.filter(User.member_id.isnot(None)).update(
        {User.member_id: None},
        synchronize_session=False,
    )
    Member.query.update({Member.referrer_id: None}, synchronize_session=False)
    MarketplaceLead.query.filter(MarketplaceLead.attributed_member_id.isnot(None)).update(
        {MarketplaceLead.attributed_member_id: None},
        synchronize_session=False,
    )

    ProductCommissionAdAllocation.query.delete(synchronize_session=False)
    ProductCommissionShare.query.delete(synchronize_session=False)
    ProductCommission.query.delete(synchronize_session=False)
    AdSplitMember.query.delete(synchronize_session=False)
    MemberLedger.query.delete(synchronize_session=False)
    PayoutNotification.query.delete(synchronize_session=False)
    OmpdFundEntry.query.delete(synchronize_session=False)
    PayoutRequest.query.delete(synchronize_session=False)
    SharingEntry.query.delete(synchronize_session=False)
    SharingBatch.query.delete(synchronize_session=False)
    ProjectBilling.query.delete(synchronize_session=False)
    ProjectCommission.query.delete(synchronize_session=False)
    Contractor.query.delete(synchronize_session=False)
    Supplier.query.delete(synchronize_session=False)
    Member.query.delete(synchronize_session=False)
    db.session.commit()


def _validate_choice(value, allowed, field_name):
    if value is None:
        return None
    if value not in allowed:
        raise ValueError(f"Invalid {field_name} '{value}'. Allowed: {', '.join(allowed)}.")
    return value


def _row_payload(row, include_lifetime_cap=False):
    member_id = int(row["member_id"])
    referrer_id = None
    if not _is_blank(row.get("referrer_id")):
        referrer_id = int(row["referrer_id"])

    status = _validate_choice(
        _clean_str(row.get("status")) or "Active",
        MEMBER_STATUSES,
        "status",
    )
    separation_type = _validate_choice(
        _clean_str(row.get("termination_type")),
        MEMBER_SEPARATION_TYPES,
        "termination_type",
    )

    payload = {
        "batch": int(row["batch"]),
        "referrer_id": referrer_id,
        "membership_type": _clean_str(row.get("membership_type")),
        "date_joined": _parse_date(row.get("date_joined")),
        "last_name": _clean_str(row.get("last_name")) or "",
        "first_name": _clean_str(row.get("first_name")) or _clean_str(row.get("last_name")) or "N/A",
        "middle_name": _clean_str(row.get("middle_name")),
        "suffix": _clean_str(row.get("suffix")),
        "address": _clean_str(row.get("address")),
        "phone": _clean_str(row.get("phone")),
        "email": _clean_str(row.get("email")),
        "birth_date": _parse_date(row.get("birth_date")),
        "gender": _clean_str(row.get("gender")),
        "civil_status": _clean_str(row.get("civil_status")),
        "highest_education": _clean_str(row.get("highest_education")),
        "occupation_income_source": _clean_str(row.get("occupation_income_source")),
        "monthly_income": _clean_str(row.get("monthly_income")),
        "number_of_dependents": _parse_dependents(row.get("number_of_dependents")),
        "beneficiary_name": _clean_str(row.get("beneficiary_name")),
        "beneficiary_address": _clean_str(row.get("beneficiary_address")),
        "beneficiary_phone": _clean_str(row.get("beneficiary_phone")),
        "beneficiary_relationship": _clean_str(row.get("beneficiary_relationship")),
        "gcash_number": _clean_str(row.get("gcash_number")),
        "bank_account_number": _clean_str(row.get("bank_account_number")),
        "bank_name": _clean_str(row.get("bank_name")),
        "age": _parse_age(row.get("age")),
        "id_picture_location": _clean_str(row.get("id_picture_location")),
        "status": status,
        "termination_date": _parse_date(row.get("termination_date")),
        "termination_type": separation_type,
    }
    if include_lifetime_cap:
        payload.update({
            "lifetime_cap_enabled": _parse_bool(row.get("lifetime_cap_enabled"), True),
            "lifetime_cap_amount": _parse_decimal(
                row.get("lifetime_cap_amount"),
                float(MEMBER_LIFETIME_EARNINGS_CAP),
            ),
        })
    return member_id, payload


def preview_members_dataframe(df, limit=5):
    rows = sorted(df.to_dict("records"), key=lambda r: (int(r["batch"]), int(r["member_id"])))
    preview = []
    for row in rows[:limit]:
        member_id, payload = _row_payload(row)
        preview.append({
            "member_id": member_id,
            "batch": payload["batch"],
            "referrer_id": payload["referrer_id"],
            "status": payload["status"],
            "date_joined": payload["date_joined"].isoformat() if payload["date_joined"] else None,
            "full_name": " ".join(
                p for p in [
                    payload["first_name"],
                    payload["middle_name"],
                    payload["last_name"],
                    payload.get("suffix"),
                ] if p
            ),
        })
    return {"row_count": len(rows), "preview": preview}


def import_members_dataframe(df, replace=False, actor_role=None):
    if replace:
        _clear_members_for_replace()

    imported = 0
    updated = 0
    skipped_referrers = []
    rows = sorted(df.to_dict("records"), key=lambda r: (int(r["batch"]), int(r["member_id"])))
    member_ids = {int(row["member_id"]) for row in rows}
    include_lifetime_cap = is_admin_role(actor_role)

    with db.session.no_autoflush:
        for row in rows:
            member_id, payload = _row_payload(row, include_lifetime_cap=include_lifetime_cap)
            referrer_id = payload["referrer_id"]
            if referrer_id is not None and referrer_id not in member_ids:
                skipped_referrers.append({"member_id": member_id, "referrer_id": referrer_id})
                payload["referrer_id"] = None

            if not payload["last_name"]:
                raise ValueError(f"Row with member_id {member_id} is missing last name.")

            member = db.session.get(Member, member_id)
            if member:
                for key, value in payload.items():
                    setattr(member, key, value)
                updated += 1
            else:
                db.session.add(Member(member_id=member_id, **payload))
                imported += 1

    db.session.commit()
    result = {"imported": imported, "updated": updated, "total": Member.query.count()}
    if skipped_referrers:
        result["skipped_referrers"] = skipped_referrers
    return result


def import_members_from_upload(file_storage, replace=False, actor_role=None):
    df = load_members_dataframe(file_storage)
    return import_members_dataframe(df, replace=replace, actor_role=actor_role)


def preview_members_upload(file_storage, limit=5):
    df = load_members_dataframe(file_storage)
    return preview_members_dataframe(df, limit=limit)


def import_members_from_xlsx(path=None, replace=False, actor_role=None):
    xlsx_path = Path(path or MEMBERS_XLSX)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Members file not found: {xlsx_path}")

    df = load_members_dataframe(xlsx_path)
    return import_members_dataframe(df, replace=replace, actor_role=actor_role)
