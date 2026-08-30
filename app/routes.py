import json
from io import BytesIO
from flask import (
    Blueprint,
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, text
from sqlalchemy.orm import joinedload
from werkzeug.security import generate_password_hash

from app import db
from app.auth import admin_required, login_required, staff_or_admin_required
from app.site_content_service import (
    get_all_ecosystem_pages,
    get_contractors,
    get_ecosystem_page,
    get_ecosystem_slugs,
    get_landing_ecosystem_section,
    get_marketplace_summary,
    apply_portal_partner_profile,
    get_partner_by_slug,
    get_portal_record_for_partner,
    partner_portal_link,
    get_suppliers,
)
from app.config import (
    CLIENT_POOL_PERCENT,
    COMMISSION_SCHEME_CLIENT,
    COMMISSION_SCHEME_CONTRACTOR,
    COMMISSION_SCHEMES,
    CONTRACTOR_POOL_PERCENT,
    CONTRACTORS_SHEET,
    SUPPLIERS_SHEET,
    DEFAULT_PRODUCT_POOL_LEVELS,
    MAX_SHARING_LEVELS,
    MEMBERS_SHEET,
    MEMBER_LIFETIME_EARNINGS_CAP,
    MEMBER_LIFETIME_PROJECT_CAP_AFTER_LIMIT,
    MEMBER_SELF_EDITABLE_FIELDS,
    MEMBER_SEPARATION_TYPES,
    MEMBER_STATUSES,
    PAYOUT_RELEASE_METHODS,
    PAYOUT_STATUS_APPROVED,
    PAYOUT_STATUS_PENDING,
    PAYOUT_STATUS_RELEASED,
    PAYOUT_STATUS_RELEASE_SUBMITTED,
    PRODUCT_AD_FUND_PERCENT,
    PRODUCT_BONUS_AD_DEFAULT_PERCENT,
    PRODUCT_BONUS_AUTO_PERCENT,
    PRODUCT_BONUS_TYPE_AD,
    PRODUCT_BONUS_TYPE_AUTO,
    PRODUCT_PLATFORM_PERCENT,
    PRODUCT_POP_PERCENT,
    PRODUCT_REF_BUYER_PERCENT,
    PRODUCT_REF_SELLER_PERCENT,
    MARKETPLACE_CATEGORIES,
    MARKETPLACE_CATEGORY_SLUGS,
    MARKETPLACE_CATEGORY_PRODUCTS,
    MARKETPLACE_FUNNEL_CATEGORY_SLUGS,
    MARKETPLACE_LEAD_STATUS_LABELS,
    MARKETPLACE_LEAD_STATUSES,
    MARKETPLACE_LEAD_ACTION_LABELS,
    MARKETPLACE_LEAD_ACTIONS,
    MARKETPLACE_LEAD_RESULT_LABELS,
    MARKETPLACE_LEAD_RESULTS,
    USER_ROLE_ADMIN,
    USER_ROLE_MEMBER,
    USER_ROLE_PORTAL_ADMIN,
    USER_ROLE_STAFF,
    assignable_user_roles,
    can_approve_payout_release,
    can_approve_payout_request,
    can_request_payout,
    can_submit_payout_release,
    can_view_payout_reports,
    can_view_payout_scheme,
    can_purge_member_database,
    can_manage_site_content,
    can_view_marketplace_help,
    can_view_features_process_flow,
    can_access_marketplace_crm,
    payout_scheme_summary,
    is_admin_role,
    is_member_role,
    is_portal_admin_role,
    is_site_admin_role,
    is_staff_role,
    normalize_role,
    post_login_redirect,
    staff_may_manage_user,
)

from app.user_manual_content import (
    APP_FEATURES,
    APP_PROCESS_FLOW,
    MARKETPLACE_CRM_GUIDE,
    resolve_user_manual,
)
from app.contractor_import_service import (
    import_contractors_from_upload,
    import_contractors_from_xlsx,
    preview_contractors_upload,
)
from app.supplier_import_service import (
    import_suppliers_from_upload,
    import_suppliers_from_xlsx,
    preview_suppliers_upload,
)
from app.hierarchy_service import (
    build_hierarchy_tree,
    build_member_hierarchy_tree,
    dashboard_stats,
    member_dashboard_stats,
    member_lineage,
)
from app.member_support_service import member_support_subjects
from app.accessibility_service import (
    get_user_accessibility_prefs,
    save_user_accessibility_prefs,
)
from app.import_service import (
    clear_members_and_dependents,
    import_members_from_upload,
    import_members_from_xlsx,
    preview_members_upload,
)
from app.models import (
    AdSplitMember,
    CommissionLevel,
    Contractor,
    MarketplaceLead,
    Member,
    MemberLedger,
    OmpdFundEntry,
    PayoutRequest,
    PayoutNotification,
    ProductCommission,
    ProductCommissionAdAllocation,
    ProductCommissionShare,
    ProjectBilling,
    ProjectCommission,
    SharingBatch,
    SharingEntry,
    Supplier,
    User,
)
from app.ledger_service import member_ledger_rows, member_ledger_stats
from app.payout_service import (
    approve_payout_release,
    approve_payout_request,
    create_payout_request,
    fund_release_rows,
    fund_release_summary,
    ompd_fund_rows,
    ompd_fund_summary,
    payout_request_query,
    payout_to_dict,
    reject_payout_release,
    reject_payout_request,
    submit_payout_release,
)
from app.platform_about import PLATFORM_ABOUT, PLATFORM_DEVELOPER
from app.permissions import (
    access_denied_response,
    forbid_member_portal_users,
    forbid_unless_member_self,
    forbid_unless_own_ledger,
    member_may_access,
    require_linked_member,
    require_marketplace_member,
)
from app.prof_reports_service import (
    commission_summary_report,
    generated_projects_index,
    project_detail_report,
)
from app.pdf_reports import (
    build_commission_summary_pdf,
    build_fund_release_pdf,
    build_project_report_pdf,
)
from app.prof_sharing_service import (
    billing_dates_with_billings,
    billings_for_billing_date,
    generate_profit_sharing,
    generated_billing_dates,
    get_commission_levels,
    remove_sharing_batch,
    sharing_batch_summary,
    ungenerated_billing_dates,
)
from app.product_commission_service import (
    compute_product_commission,
    delete_product_commission,
    save_product_commission,
)
from app.marketplace_service import (
    create_lead,
    ensure_member_share_code,
    get_attributed_member,
    get_marketplace_category,
    get_member_by_share_code,
    get_published_listing,
    landing_featured_counts,
    list_published_by_category,
    listing_crm_stats,
    marketplace_crm_overview,
    marketplace_listing_options,
    member_lead_count,
    member_leads,
    products_page_content,
    funnel_page_content,
    search_marketplace_leads,
    serialize_lead_history,
    set_attribution_cookie,
    clear_attribution_cookie,
    update_marketplace_lead,
    update_marketplace_lead_status,
    get_marketplace_lead_history,
)

main_routes = Blueprint("main_routes", __name__)

MEMBER_TEMPLATE_COLUMNS = [
    "member_id",
    "batch",
    "referrer_id",
    "membership_type",
    "date_joined",
    "last_name",
    "first_name",
    "middle_name",
    "suffix",
    "address",
    "phone",
    "email",
    "birth_date",
    "gender",
    "civil_status",
    "highest_education",
    "occupation_income_source",
    "monthly_income",
    "number_of_dependents",
    "beneficiary_name",
    "beneficiary_address",
    "beneficiary_phone",
    "status",
    "termination_date",
    "termination_type",
    "lifetime_cap_enabled",
    "lifetime_cap_amount",
    "gcash_number",
    "bank_account_number",
    "bank_name",
    "age",
    "id_picture_location",
    "beneficiary_relationship",
]

CONTRACTOR_TEMPLATE_COLUMNS = [
    "contractor_id",
    "batch",
    "member_referrer_id",
    "company_name",
    "company_address",
    "representative_name",
    "contact_no",
    "date_joined",
]

SUPPLIER_TEMPLATE_COLUMNS = [
    "supplier_id",
    "batch",
    "member_referrer_id",
    "company_name",
    "company_address",
    "representative_name",
    "contact_no",
    "date_joined",
]


def _blank_import_template(sheet_name, columns, reference_rows=None):
    buffer = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(columns)
    worksheet.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="F0A500")
    header_font = Font(bold=True, color="0D0F12")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for index, column in enumerate(columns, start=1):
        width = min(max(len(column) + 4, 14), 32)
        worksheet.column_dimensions[get_column_letter(index)].width = width

    if reference_rows:
        reference_sheet = workbook.create_sheet("reference")
        for row in reference_rows:
            reference_sheet.append(row)
        for cell in reference_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        reference_sheet.column_dimensions["A"].width = 24
        reference_sheet.column_dimensions["B"].width = 42

    workbook.save(buffer)
    buffer.seek(0)
    return buffer


@main_routes.route("/")
def index():
    return render_template(
        "index.html",
        landing_section=get_landing_ecosystem_section(),
        ecosystem_pages=get_all_ecosystem_pages(),
        ecosystem_slugs=get_ecosystem_slugs(),
        marketplace_categories=MARKETPLACE_CATEGORIES,
        marketplace_category_slugs=MARKETPLACE_CATEGORY_SLUGS,
        marketplace_counts=landing_featured_counts(),
    )


@main_routes.route("/ecosystem/<slug>")
def ecosystem_page(slug):
    page = get_ecosystem_page(slug)
    if page is None:
        abort(404)
    all_pages = get_all_ecosystem_pages()
    related_pages = [all_pages[s] for s in get_ecosystem_slugs() if s != slug]
    context = {
        "page": page,
        "related_pages": related_pages,
        "landing_nav_active": None,
    }
    if slug == "partners":
        context["contractors"] = get_contractors()
        context["suppliers"] = get_suppliers()
    return render_template("ecosystem_page.html", **context)


def _marketplace_page_context(category, share_member=None):
    meta = get_marketplace_category(category)
    if not meta:
        abort(404)
    attributed = share_member or get_attributed_member()
    return {
        "category": category,
        "category_meta": meta,
        "categories": MARKETPLACE_CATEGORIES,
        "category_slugs": MARKETPLACE_CATEGORY_SLUGS,
        "attributed_member": attributed,
        "share_member": share_member,
        "executive_summary": get_marketplace_summary(category),
        "landing_nav_active": None,
    }


@main_routes.route("/marketplace/<category>", methods=["GET", "POST"])
def marketplace_category(category):
    ctx = _marketplace_page_context(category)
    ctx["listings"] = list_published_by_category(category)

    if category in MARKETPLACE_FUNNEL_CATEGORY_SLUGS:
        ctx["page"] = funnel_page_content(category)
        return render_template("marketplace_funnel.html", **ctx)

    return render_template("marketplace_category.html", **ctx)


@main_routes.route("/marketplace/<category>/<int:listing_id>", methods=["GET", "POST"])
def marketplace_detail(category, listing_id):
    listing = get_published_listing(category, listing_id)
    if not listing:
        abort(404)
    ctx = _marketplace_page_context(category)
    ctx["listing"] = listing
    form_error = None
    form_success = None
    form_values = {"guest_name": "", "guest_phone": "", "guest_email": "", "message": ""}

    if request.method == "POST":
        form_values = {
            "guest_name": request.form.get("guest_name") or "",
            "guest_phone": request.form.get("guest_phone") or "",
            "guest_email": request.form.get("guest_email") or "",
            "message": request.form.get("message") or "",
        }
        try:
            create_lead(
                listing,
                guest_name=form_values["guest_name"],
                guest_phone=form_values["guest_phone"],
                guest_email=form_values["guest_email"],
                message=form_values["message"],
                source_path=request.path,
            )
            form_success = "Thank you. Your inquiry was sent. A TBGP representative will follow up."
            form_values = {"guest_name": "", "guest_phone": "", "guest_email": "", "message": ""}
        except ValueError as exc:
            form_error = str(exc)

    ctx["form_error"] = form_error
    ctx["form_success"] = form_success
    ctx["form_values"] = form_values
    return render_template("marketplace_detail.html", **ctx)


@main_routes.route("/m/<share_code>/marketplace/<category>")
def marketplace_category_shared(share_code, category):
    member = get_member_by_share_code(share_code)
    if not member:
        abort(404)
    response = redirect(url_for("main_routes.marketplace_category", category=category))
    return set_attribution_cookie(response, member)


@main_routes.route("/m/<share_code>/marketplace/<category>/<int:listing_id>")
def marketplace_detail_shared(share_code, category, listing_id):
    member = get_member_by_share_code(share_code)
    if not member:
        abort(404)
    response = redirect(
        url_for("main_routes.marketplace_detail", category=category, listing_id=listing_id)
    )
    return set_attribution_cookie(response, member)


@main_routes.route("/m/<share_code>/marketplace")
def marketplace_hub_shared(share_code):
    member = get_member_by_share_code(share_code)
    if not member:
        abort(404)
    response = redirect(
        url_for("main_routes.marketplace_category", category=MARKETPLACE_CATEGORY_PRODUCTS)
    )
    return set_attribution_cookie(response, member)


@main_routes.route("/my-marketplace")
@login_required
def my_marketplace():
    user = User.query.get(session.get("user_id"))
    if not (is_member_role(user.role) or is_staff_role(user.role)):
        return redirect(url_for("main_routes.members"))

    linked_id = require_marketplace_member()
    if not linked_id:
        return access_denied_response(
            "My Marketplace requires a linked Member ID on your account. "
            "Ask an Admin to link your user to a member record."
        )

    member = db.session.get(Member, linked_id)
    if not member:
        return access_denied_response(
            "Your linked member record was not found. Ask an Admin to update your account."
        )

    share_code = ensure_member_share_code(member)
    base_url = request.url_root.rstrip("/")
    share_links = {
        "hub": f"{base_url}/m/{share_code}/marketplace",
    }
    for slug in MARKETPLACE_CATEGORY_SLUGS:
        share_links[slug] = f"{base_url}/m/{share_code}/marketplace/{slug}"

    leads = member_leads(linked_id)
    return render_template(
        "my_marketplace.html",
        fullname=user.full_name or ("Member" if is_member_role(user.role) else "Staff"),
        role=normalize_role(user.role),
        active_page="my_marketplace",
        member=member,
        share_code=share_code,
        share_links=share_links,
        categories=MARKETPLACE_CATEGORIES,
        category_slugs=MARKETPLACE_CATEGORY_SLUGS,
        leads=leads,
        lead_count=member_lead_count(linked_id),
        lead_status_labels=MARKETPLACE_LEAD_STATUS_LABELS,
        lead_action_labels=MARKETPLACE_LEAD_ACTION_LABELS,
        lead_result_labels=MARKETPLACE_LEAD_RESULT_LABELS,
    )


def _marketplace_crm_filters_from_request():
    referrer_raw = (request.args.get("referrer_id") or "").strip()
    listing_raw = (request.args.get("listing_id") or "").strip()
    referrer_id = None
    listing_id = None
    if referrer_raw.isdigit():
        referrer_id = int(referrer_raw)
    if listing_raw.isdigit():
        listing_id = int(listing_raw)
    return {
        "q": (request.args.get("q") or "").strip(),
        "category": (request.args.get("category") or "").strip() or None,
        "listing_id": listing_id,
        "referrer_id": referrer_id,
        "attribution": (request.args.get("attribution") or "").strip() or None,
        "status": (request.args.get("status") or "").strip() or None,
        "date_from": (request.args.get("date_from") or "").strip() or None,
        "date_to": (request.args.get("date_to") or "").strip() or None,
    }


@main_routes.route("/admin/marketplace-crm")
@login_required
def marketplace_crm():
    user = _dashboard_user()
    if not can_access_marketplace_crm(user.role):
        return access_denied_response(
            "Marketplace CRM is available to Admin, SiteAdmin, and Staff roles."
        )

    filters = _marketplace_crm_filters_from_request()
    if filters["category"] and filters["category"] not in MARKETPLACE_CATEGORIES:
        filters["category"] = None
    if filters["status"] and filters["status"] not in MARKETPLACE_LEAD_STATUSES:
        filters["status"] = None

    overview = marketplace_crm_overview()
    leads = search_marketplace_leads(**filters)
    listing_options = marketplace_listing_options(filters["category"])

    return render_template(
        "marketplace_crm.html",
        fullname=user.full_name or "Admin",
        role=normalize_role(user.role),
        active_page="marketplace_crm",
        overview=overview,
        leads=leads,
        filters=filters,
        categories=MARKETPLACE_CATEGORIES,
        category_slugs=MARKETPLACE_CATEGORY_SLUGS,
        listing_options=listing_options,
        lead_statuses=MARKETPLACE_LEAD_STATUSES,
        lead_status_labels=MARKETPLACE_LEAD_STATUS_LABELS,
        lead_actions=MARKETPLACE_LEAD_ACTIONS,
        lead_action_labels=MARKETPLACE_LEAD_ACTION_LABELS,
        lead_results=MARKETPLACE_LEAD_RESULTS,
        lead_result_labels=MARKETPLACE_LEAD_RESULT_LABELS,
        can_edit_listings=can_manage_site_content(user.role),
    )


@main_routes.route("/admin/marketplace-crm/listing/<int:listing_id>")
@login_required
def marketplace_crm_listing(listing_id):
    user = _dashboard_user()
    if not can_access_marketplace_crm(user.role):
        return access_denied_response(
            "Marketplace CRM is available to Admin, SiteAdmin, and Staff roles."
        )

    stats = listing_crm_stats(listing_id)
    if not stats:
        abort(404)

    return render_template(
        "marketplace_crm_listing.html",
        fullname=user.full_name or "Admin",
        role=normalize_role(user.role),
        active_page="marketplace_crm",
        stats=stats,
        listing=stats["listing"],
        categories=MARKETPLACE_CATEGORIES,
        category_slugs=MARKETPLACE_CATEGORY_SLUGS,
        lead_statuses=MARKETPLACE_LEAD_STATUSES,
        lead_status_labels=MARKETPLACE_LEAD_STATUS_LABELS,
        lead_actions=MARKETPLACE_LEAD_ACTIONS,
        lead_action_labels=MARKETPLACE_LEAD_ACTION_LABELS,
        lead_results=MARKETPLACE_LEAD_RESULTS,
        lead_result_labels=MARKETPLACE_LEAD_RESULT_LABELS,
        can_edit_listings=can_manage_site_content(user.role),
    )


@main_routes.route("/admin/marketplace-crm/leads/<int:lead_id>/status", methods=["POST"])
@login_required
def marketplace_crm_lead_status(lead_id):
    user = _dashboard_user()
    if not can_access_marketplace_crm(user.role):
        return access_denied_response(
            "Marketplace CRM is available to Admin, SiteAdmin, and Staff roles."
        )

    data = request.get_json(silent=True) if request.is_json else request.form
    data = data or {}
    status = data.get("status")
    action_required = data.get("action_required")
    final_result = data.get("final_result")
    note = data.get("note")
    try:
        lead = update_marketplace_lead(
            lead_id,
            status=status if status is not None else None,
            action_required=action_required if "action_required" in data else None,
            final_result=final_result if "final_result" in data else None,
            note=note,
            created_by_user_id=user.user_id,
        )
        return jsonify({
            "status": "success",
            "msg": f"Inquiry #{lead.lead_id} updated.",
            "lead_id": lead.lead_id,
            "lead_status": lead.status,
            "lead_status_label": MARKETPLACE_LEAD_STATUS_LABELS.get(lead.status, lead.status),
            "action_required": lead.action_required or "",
            "action_required_label": MARKETPLACE_LEAD_ACTION_LABELS.get(lead.action_required or "", ""),
            "final_result": lead.final_result or "",
            "final_result_label": MARKETPLACE_LEAD_RESULT_LABELS.get(lead.final_result or "", ""),
            "aging_days": lead.aging_days,
        })
    except ValueError as exc:
        return jsonify({"status": "error", "msg": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Could not update inquiry: {exc}"}), 500


@main_routes.route("/admin/marketplace-crm/leads/<int:lead_id>/history")
@login_required
def marketplace_crm_lead_history(lead_id):
    user = _dashboard_user()
    if not can_access_marketplace_crm(user.role):
        return access_denied_response(
            "Marketplace CRM is available to Admin, SiteAdmin, and Staff roles."
        )

    lead = db.session.get(MarketplaceLead, lead_id)
    if not lead:
        return jsonify({"status": "error", "msg": "Inquiry not found."}), 404

    entries = get_marketplace_lead_history(lead_id)
    return jsonify({
        "status": "success",
        "lead_id": lead_id,
        "aging_days": lead.aging_days,
        "history": serialize_lead_history(entries),
    })


@main_routes.route("/admin/marketplace-crm/export.csv")
@login_required
def marketplace_crm_export():
    user = _dashboard_user()
    if not can_access_marketplace_crm(user.role):
        return access_denied_response(
            "Marketplace CRM is available to Admin, SiteAdmin, and Staff roles."
        )

    filters = _marketplace_crm_filters_from_request()
    if filters["category"] and filters["category"] not in MARKETPLACE_CATEGORIES:
        filters["category"] = None
    if filters["status"] and filters["status"] not in MARKETPLACE_LEAD_STATUSES:
        filters["status"] = None
    leads = search_marketplace_leads(**filters, limit=5000)

    output = BytesIO()
    # UTF-8 BOM for Excel
    output.write("\ufeff".encode("utf-8"))
    writer_lines = [
        "lead_id,created_at,aging_days,status,action_required,final_result,category,listing_id,listing_title,"
        "guest_name,guest_phone,guest_email,message,referred_by_id,referred_by_name,source_path\n"
    ]
    for lead in leads:
        listing = lead.listing
        member = lead.attributed_member
        category = listing.category if listing else (lead.interest_category or "")
        category_label = MARKETPLACE_CATEGORIES.get(category, {}).get("label", category)
        cells = [
            lead.lead_id,
            lead.created_at.isoformat(sep=" ", timespec="minutes") if lead.created_at else "",
            lead.aging_days,
            MARKETPLACE_LEAD_STATUS_LABELS.get(lead.status or "new", lead.status or "new"),
            MARKETPLACE_LEAD_ACTION_LABELS.get(lead.action_required or "", lead.action_required or ""),
            MARKETPLACE_LEAD_RESULT_LABELS.get(lead.final_result or "", lead.final_result or ""),
            category_label,
            lead.listing_id or "",
            (listing.title if listing else "").replace('"', '""'),
            (lead.guest_name or "").replace('"', '""'),
            lead.guest_phone or "",
            lead.guest_email or "",
            (lead.message or "").replace('"', '""').replace("\n", " "),
            lead.attributed_member_id or "",
            (member.full_name if member else "").replace('"', '""'),
            lead.source_path or "",
        ]
        writer_lines.append(
            ",".join(f'"{cell}"' if isinstance(cell, str) else str(cell) for cell in cells) + "\n"
        )
    output.write("".join(writer_lines).encode("utf-8"))
    output.seek(0)
    return send_file(
        output,
        mimetype="text/csv",
        as_attachment=True,
        download_name=f"marketplace-crm-{datetime.utcnow().strftime('%Y%m%d-%H%M')}.csv",
    )


@main_routes.route("/partners/<partner_slug>")
def partner_profile(partner_slug):
    partner = get_partner_by_slug(partner_slug)
    if partner is None:
        abort(404)
    portal_record = get_portal_record_for_partner(partner)
    link = partner_portal_link(partner)
    partner = apply_portal_partner_profile(partner, portal_record)
    return render_template(
        "partner_profile.html",
        partner=partner,
        portal_contractor=link["portal_contractor"],
        portal_supplier=link["portal_supplier"],
        member_referrer=link["member_referrer"],
        landing_nav_active=None,
    )


@main_routes.route("/logout")
def logout():
    session.clear()
    response = redirect(url_for("main_routes.index"))
    return clear_attribution_cookie(response)


@main_routes.route("/api/accessibility-preferences", methods=["GET", "POST"])
@login_required
def accessibility_preferences():
    user_id = session.get("user_id")
    if request.method == "GET":
        return jsonify(get_user_accessibility_prefs(user_id))

    payload = request.get_json(silent=True) or request.form
    prefs = save_user_accessibility_prefs(user_id, payload)
    return jsonify(prefs)


@main_routes.route("/login", methods=["GET", "POST"])
def login():
    next_param = request.args.get("next", "")

    if request.method == "POST":
        data = request.get_json() if request.is_json else request.form
        username = data.get("username", "").strip()
        password = data.get("password", "").strip()
        next_param = data.get("next", next_param)

        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            session["user_id"] = user.user_id
            session["username"] = user.username
            session["fullname"] = user.full_name
            session["role"] = normalize_role(user.role)
            session["member_id"] = user.member_id

            if is_member_role(user.role) and not user.member_id:
                message = "Member login is not linked to a member record. Contact an administrator."
                if request.is_json:
                    return jsonify({"success": False, "message": message})
                flash(message, "danger")
                session.clear()
                return render_template("login.html", next_url=next_param)

            redirect_to = post_login_redirect(user.role, next_param)

            if request.is_json:
                return jsonify({"success": True, "redirect": redirect_to})
            flash("Login successful", "success")
            return redirect(redirect_to)

        message = "Invalid credentials"
        if request.is_json:
            return jsonify({"success": False, "message": message})
        flash(message, "danger")

    return render_template("login.html", next_url=next_param)


@main_routes.route("/dashboard")
@login_required
def dashboard():
    user = User.query.get(session.get("user_id"))
    if not user:
        return redirect(url_for("main_routes.logout"))

    if is_site_admin_role(user.role):
        return redirect(url_for("site_admin.home"))

    if is_member_role(user.role):
        linked_id = require_linked_member()
        personal_stats = member_dashboard_stats(linked_id) if linked_id else None
        return render_template(
            "dash.html",
            fullname=user.full_name or "Member",
            role=normalize_role(user.role),
            stats=None,
            personal_stats=personal_stats,
            is_personal_dashboard=True,
            member_support_subjects=member_support_subjects(),
            active_page="dashboard",
        )

    stats = dashboard_stats()
    return render_template(
        "dash.html",
        fullname=user.full_name or "Admin",
        role=normalize_role(user.role),
        stats=stats,
        personal_stats=None,
        is_personal_dashboard=False,
        active_page="dashboard",
    )


@main_routes.route("/help/user-manual")
@login_required
def user_manual():
    user = _dashboard_user()
    role = normalize_role(user.role)
    if is_site_admin_role(user.role) and not is_portal_admin_role(user.role):
        return redirect(url_for("site_admin.user_manual"))
    manual, manual_role, manual_choices = resolve_user_manual(
        user.role,
        request.args.get("manual"),
    )
    return render_template(
        "user_manual.html",
        fullname=user.full_name or "User",
        role=role,
        manual_role=manual_role,
        manual_choices=manual_choices,
        manual_url_endpoint="main_routes.user_manual",
        active_page="user_manual",
        manual=manual,
    )


@main_routes.route("/about/features-process-flow")
@login_required
def about_features_process_flow():
    user = _dashboard_user()
    if not can_view_features_process_flow(user.role):
        return access_denied_response(
            "App Features & Process Flow is available to Admin, Staff, and SiteAdmin roles."
        )
    role = normalize_role(user.role)
    return render_template(
        "about_features_process_flow.html",
        fullname=user.full_name or "User",
        role=role,
        active_page="about_features_process_flow",
        features=APP_FEATURES,
        process_flow=APP_PROCESS_FLOW,
    )


@main_routes.route("/help/marketplace-crm")
@login_required
def help_marketplace_crm():
    user = _dashboard_user()
    if not can_view_marketplace_help(user.role):
        return access_denied_response(
            "The Marketplace & CRM guide is available to Admin, SiteAdmin, and Staff roles."
        )
    role = normalize_role(user.role)
    return render_template(
        "help_marketplace_crm.html",
        fullname=user.full_name or "User",
        role=role,
        active_page="help_marketplace_crm",
        guide=MARKETPLACE_CRM_GUIDE,
    )


@main_routes.route("/about/platform")
@login_required
def about_platform():
    user = _dashboard_user()
    role = normalize_role(user.role)
    return render_template(
        "about_platform.html",
        fullname=user.full_name or "User",
        role=role,
        active_page="about_platform",
        platform=PLATFORM_ABOUT,
        developer=PLATFORM_DEVELOPER,
    )


@main_routes.route("/members")
@login_required
def members():
    user = User.query.get(session.get("user_id"))
    query = (
        Member.query
        .options(joinedload(Member.referrer), joinedload(Member.referrals))
        .order_by(Member.batch.asc(), Member.member_id.asc())
    )
    if is_member_role(user.role):
        linked_id = require_linked_member()
        if not linked_id:
            return access_denied_response("Your account is not linked to a member record.")
        query = query.filter(Member.member_id == linked_id)

    members_list = query.all()
    return render_template(
        "members.html",
        fullname=user.full_name or "Admin",
        role=normalize_role(user.role),
        members=members_list,
        active_page="members",
        is_own_profile_view=is_member_role(user.role),
    )


@main_routes.route("/members/my-profile", methods=["POST"])
@login_required
def member_update_own_profile():
    if not is_member_role(session.get("role")):
        return access_denied_response("Access denied.")

    linked_id = require_linked_member()
    if not linked_id:
        return access_denied_response("Your account is not linked to a member record.")

    member = db.session.get(Member, linked_id)
    if not member:
        return jsonify({"status": "error", "msg": "Member not found."}), 404

    data = request.get_json() if request.is_json else request.form

    try:
        _apply_member_self_form(member, data)
        db.session.commit()
        return jsonify({
            "status": "success",
            "msg": "Your profile has been updated.",
            "member": member.to_dict(),
        })
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Database error: {exc}"}), 500


@main_routes.route("/members/my-password", methods=["POST"])
@login_required
def member_change_own_password():
    if not is_member_role(session.get("role")):
        return access_denied_response("Access denied.")

    if not require_linked_member():
        return access_denied_response("Your account is not linked to a member record.")

    user = db.session.get(User, session.get("user_id"))
    if not user:
        return jsonify({"status": "error", "msg": "User not found."}), 404

    data = request.get_json(silent=True) if request.is_json else request.form
    current_password = (data.get("current_password") if data else "") or ""
    current_password = current_password.strip()
    new_password = (data.get("new_password") if data else "") or ""
    new_password = new_password.strip()
    confirm_password = (data.get("confirm_password") if data else "") or ""
    confirm_password = confirm_password.strip()

    if not current_password or not new_password or not confirm_password:
        return jsonify({"status": "error", "msg": "All password fields are required."}), 400

    if not user.check_password(current_password):
        return jsonify({"status": "error", "msg": "Current password is incorrect."}), 400

    if new_password != confirm_password:
        return jsonify({"status": "error", "msg": "New password and confirmation do not match."}), 400

    if len(new_password) < 4:
        return jsonify({"status": "error", "msg": "New password must be at least 4 characters."}), 400

    if user.check_password(new_password):
        return jsonify({"status": "error", "msg": "New password must be different from your current password."}), 400

    try:
        user.set_password(new_password)
        db.session.commit()
        return jsonify({"status": "success", "msg": "Your password has been updated."})
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Database error: {exc}"}), 500


@main_routes.route("/members/template")
@login_required
@staff_or_admin_required
def members_template():
    workbook = _blank_import_template(
        MEMBERS_SHEET,
        MEMBER_TEMPLATE_COLUMNS,
        reference_rows=[
            ["field", "allowed values / note"],
            ["status", ", ".join(MEMBER_STATUSES)],
            ["termination_type", ", ".join(MEMBER_SEPARATION_TYPES)],
            ["date fields", "Use YYYY-MM-DD format."],
            ["referrer_id", "Leave blank only for batch 1 root members."],
            ["lifetime_cap_enabled", "Admin-only. Use true/false. Blank defaults to true."],
            ["lifetime_cap_amount", f"Admin-only threshold. Blank defaults to {MEMBER_LIFETIME_EARNINGS_CAP:,.2f}."],
            [
                "lifetime limit rule",
                f"After the threshold, project earnings are capped at {MEMBER_LIFETIME_PROJECT_CAP_AFTER_LIMIT:,.2f}; excess goes to POP Lifetime Limit Fund.",
            ],
        ],
    )
    return send_file(
        workbook,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="members_import_template.xlsx",
    )


@main_routes.route("/members/preview", methods=["POST"])
@login_required
@staff_or_admin_required
def preview_members():
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"status": "error", "msg": "Please choose an Excel file."}), 400
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"status": "error", "msg": "Only .xlsx or .xls files are supported."}), 400

    try:
        result = preview_members_upload(file)
        return jsonify({"status": "success", **result})
    except ValueError as exc:
        return jsonify({"status": "error", "msg": str(exc)}), 400
    except Exception as exc:
        return jsonify({"status": "error", "msg": f"Could not parse file: {exc}"}), 500


@main_routes.route("/members/add", methods=["POST"])
@login_required
@staff_or_admin_required
def add_member():
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"status": "error", "msg": "Please choose an Excel file."}), 400
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"status": "error", "msg": "Only .xlsx or .xls files are supported."}), 400

    try:
        result = import_members_from_upload(file, replace=False, actor_role=session.get("role"))
        return jsonify({
            "status": "success",
            "msg": (
                f"Import complete: {result['imported']} added, "
                f"{result['updated']} updated ({result['total']} total members)."
            ),
            **result,
        })
    except ValueError as exc:
        return jsonify({"status": "error", "msg": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Import failed: {exc}"}), 500


@main_routes.route("/contractors")
@login_required
def contractors():
    denied = forbid_member_portal_users()
    if denied:
        return denied

    user = User.query.get(session.get("user_id"))
    contractors_list = (
        Contractor.query
        .options(joinedload(Contractor.member_referrer))
        .order_by(Contractor.batch.asc(), Contractor.contractor_id.asc())
        .all()
    )
    members_list = Member.query.order_by(Member.batch.asc(), Member.member_id.asc()).all()
    return render_template(
        "contractors.html",
        fullname=user.full_name or "Admin",
        role=normalize_role(user.role),
        contractors=contractors_list,
        members=members_list,
        active_page="contractors",
        is_member_referral_view=False,
    )


@main_routes.route("/my-referrals/members")
@login_required
def my_member_referrals():
    user = User.query.get(session.get("user_id"))
    if not is_member_role(user.role):
        return redirect(url_for("main_routes.members"))

    linked_id = require_linked_member()
    if not linked_id:
        return access_denied_response("Your account is not linked to a member record.")

    member = Member.query.options(joinedload(Member.referrals)).get(linked_id)
    referrals = sorted(member.referrals, key=lambda row: row.member_id) if member else []
    return render_template(
        "my_member_referrals.html",
        fullname=user.full_name or "Member",
        role=normalize_role(user.role),
        referrals=referrals,
        active_page="my_member_referrals",
    )


@main_routes.route("/my-referrals/contractors")
@login_required
def my_contractor_referrals():
    user = User.query.get(session.get("user_id"))
    if not is_member_role(user.role):
        return redirect(url_for("main_routes.contractors"))

    linked_id = require_linked_member()
    if not linked_id:
        return access_denied_response("Your account is not linked to a member record.")

    contractors_list = (
        Contractor.query
        .options(joinedload(Contractor.member_referrer))
        .filter_by(member_referrer_id=linked_id)
        .order_by(Contractor.batch.asc(), Contractor.contractor_id.asc())
        .all()
    )
    return render_template(
        "contractors.html",
        fullname=user.full_name or "Member",
        role=normalize_role(user.role),
        contractors=contractors_list,
        members=[],
        active_page="my_contractor_referrals",
        is_member_referral_view=True,
    )


@main_routes.route("/my-referrals/suppliers")
@login_required
def my_supplier_referrals():
    user = User.query.get(session.get("user_id"))
    if not is_member_role(user.role):
        return redirect(url_for("main_routes.suppliers"))

    linked_id = require_linked_member()
    if not linked_id:
        return access_denied_response("Your account is not linked to a member record.")

    suppliers_list = (
        Supplier.query
        .options(joinedload(Supplier.member_referrer))
        .filter_by(member_referrer_id=linked_id)
        .order_by(Supplier.batch.asc(), Supplier.supplier_id.asc())
        .all()
    )
    return render_template(
        "suppliers.html",
        fullname=user.full_name or "Member",
        role=normalize_role(user.role),
        suppliers=suppliers_list,
        members=[],
        active_page="my_supplier_referrals",
        is_member_referral_view=True,
    )


@main_routes.route("/contractors/template")
@login_required
@staff_or_admin_required
def contractors_template():
    workbook = _blank_import_template(
        CONTRACTORS_SHEET,
        CONTRACTOR_TEMPLATE_COLUMNS,
        reference_rows=[
            ["field", "allowed values / note"],
            ["member_referrer_id", "Must match an existing member_id."],
            ["date_joined", "Use YYYY-MM-DD format."],
        ],
    )
    return send_file(
        workbook,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="contractors_import_template.xlsx",
    )


@main_routes.route("/contractors/preview", methods=["POST"])
@login_required
@staff_or_admin_required
def preview_contractors():
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"status": "error", "msg": "Please choose an Excel file."}), 400
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"status": "error", "msg": "Only .xlsx or .xls files are supported."}), 400

    try:
        result = preview_contractors_upload(file)
        return jsonify({"status": "success", **result})
    except ValueError as exc:
        return jsonify({"status": "error", "msg": str(exc)}), 400
    except Exception as exc:
        return jsonify({"status": "error", "msg": f"Could not parse file: {exc}"}), 500


@main_routes.route("/contractors/add", methods=["POST"])
@login_required
@staff_or_admin_required
def add_contractor():
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"status": "error", "msg": "Please choose an Excel file."}), 400
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"status": "error", "msg": "Only .xlsx or .xls files are supported."}), 400

    try:
        result = import_contractors_from_upload(file, replace=False)
        return jsonify({
            "status": "success",
            "msg": (
                f"Import complete: {result['imported']} added, "
                f"{result['updated']} updated ({result['total']} total contractors)."
            ),
            **result,
        })
    except ValueError as exc:
        return jsonify({"status": "error", "msg": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Import failed: {exc}"}), 500


def _apply_contractor_form(contractor, data, include_batch=False):
    company_name = (data.get("company_name") or "").strip()
    if not company_name:
        raise ValueError("Company name is required.")

    if include_batch:
        batch = _parse_form_int(data.get("batch"), "batch")
        if batch is None or batch <= 0:
            raise ValueError("Batch is required and must be greater than zero.")
        contractor.batch = batch

    referrer_raw = (data.get("member_referrer_id") or "").strip()
    if not referrer_raw:
        raise ValueError("Member referrer is required.")

    member_referrer_id = _parse_form_int(referrer_raw, "member_referrer_id")
    referrer = db.session.get(Member, member_referrer_id)
    if not referrer:
        raise ValueError("Member referrer not found.")

    contractor.company_name = company_name
    contractor.company_address = (data.get("company_address") or "").strip() or None
    contractor.representative_name = (data.get("representative_name") or "").strip() or None
    contractor.contact_no = (data.get("contact_no") or "").strip() or None
    contractor.member_referrer_id = member_referrer_id
    contractor.date_joined = _parse_form_date(data.get("date_joined"), "date_joined") or contractor.date_joined


@main_routes.route("/admin/contractors", methods=["POST"])
@login_required
@staff_or_admin_required
def admin_create_contractor():
    data = request.get_json() if request.is_json else request.form

    try:
        contractor_id = _parse_form_int(data.get("contractor_id"), "contractor_id")
        if contractor_id is None or contractor_id <= 0:
            raise ValueError("Contractor ID is required and must be greater than zero.")
        if db.session.get(Contractor, contractor_id):
            raise ValueError(f"Contractor #{contractor_id} already exists.")

        contractor = Contractor(contractor_id=contractor_id, batch=1, member_referrer_id=1, company_name="")
        _apply_contractor_form(contractor, data, include_batch=True)
        db.session.add(contractor)
        db.session.commit()
        return jsonify({
            "status": "success",
            "msg": f"Contractor #{contractor.contractor_id} added.",
            "contractor": contractor.to_dict(),
        })
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Database error: {exc}"}), 500


@main_routes.route("/admin/contractors/<int:contractor_id>", methods=["POST"])
@login_required
@staff_or_admin_required
def admin_update_contractor(contractor_id):
    contractor = db.session.get(Contractor, contractor_id)
    if not contractor:
        return jsonify({"status": "error", "msg": "Contractor not found."}), 404

    data = request.get_json() if request.is_json else request.form

    try:
        _apply_contractor_form(contractor, data)
        db.session.commit()
        return jsonify({
            "status": "success",
            "msg": f"Contractor #{contractor.contractor_id} updated.",
            "contractor": contractor.to_dict(),
        })
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Database error: {exc}"}), 500


@main_routes.route("/admin/contractors/<int:contractor_id>", methods=["DELETE"])
@login_required
@staff_or_admin_required
def admin_delete_contractor(contractor_id):
    contractor = db.session.get(Contractor, contractor_id)
    if not contractor:
        return jsonify({"status": "error", "msg": "Contractor not found."}), 404

    try:
        db.session.delete(contractor)
        db.session.commit()
        return jsonify({"status": "success", "msg": f"Contractor #{contractor_id} deleted."})
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Database error: {exc}"}), 500


@main_routes.route("/suppliers")
@login_required
def suppliers():
    denied = forbid_member_portal_users()
    if denied:
        return denied

    user = User.query.get(session.get("user_id"))
    suppliers_list = (
        Supplier.query
        .options(joinedload(Supplier.member_referrer))
        .order_by(Supplier.batch.asc(), Supplier.supplier_id.asc())
        .all()
    )
    members_list = Member.query.order_by(Member.batch.asc(), Member.member_id.asc()).all()
    return render_template(
        "suppliers.html",
        fullname=user.full_name or "Admin",
        role=normalize_role(user.role),
        suppliers=suppliers_list,
        members=members_list,
        active_page="suppliers",
        is_member_referral_view=False,
    )


@main_routes.route("/suppliers/template")
@login_required
@staff_or_admin_required
def suppliers_template():
    workbook = _blank_import_template(
        SUPPLIERS_SHEET,
        SUPPLIER_TEMPLATE_COLUMNS,
        reference_rows=[
            ["field", "allowed values / note"],
            ["member_referrer_id", "Must match an existing member_id."],
            ["date_joined", "Use YYYY-MM-DD format."],
        ],
    )
    return send_file(
        workbook,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="suppliers_import_template.xlsx",
    )


@main_routes.route("/suppliers/preview", methods=["POST"])
@login_required
@staff_or_admin_required
def preview_suppliers():
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"status": "error", "msg": "Please choose an Excel file."}), 400
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"status": "error", "msg": "Only .xlsx or .xls files are supported."}), 400

    try:
        result = preview_suppliers_upload(file)
        return jsonify({"status": "success", **result})
    except ValueError as exc:
        return jsonify({"status": "error", "msg": str(exc)}), 400
    except Exception as exc:
        return jsonify({"status": "error", "msg": f"Could not parse file: {exc}"}), 500


@main_routes.route("/suppliers/add", methods=["POST"])
@login_required
@staff_or_admin_required
def add_supplier():
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"status": "error", "msg": "Please choose an Excel file."}), 400
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"status": "error", "msg": "Only .xlsx or .xls files are supported."}), 400

    try:
        result = import_suppliers_from_upload(file, replace=False)
        return jsonify({
            "status": "success",
            "msg": (
                f"Import complete: {result['imported']} added, "
                f"{result['updated']} updated ({result['total']} total suppliers)."
            ),
            **result,
        })
    except ValueError as exc:
        return jsonify({"status": "error", "msg": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Import failed: {exc}"}), 500


def _apply_supplier_form(supplier, data, include_batch=False):
    company_name = (data.get("company_name") or "").strip()
    if not company_name:
        raise ValueError("Company name is required.")

    if include_batch:
        batch = _parse_form_int(data.get("batch"), "batch")
        if batch is None or batch <= 0:
            raise ValueError("Batch is required and must be greater than zero.")
        supplier.batch = batch

    referrer_raw = (data.get("member_referrer_id") or "").strip()
    if not referrer_raw:
        raise ValueError("Member referrer is required.")

    member_referrer_id = _parse_form_int(referrer_raw, "member_referrer_id")
    referrer = db.session.get(Member, member_referrer_id)
    if not referrer:
        raise ValueError("Member referrer not found.")

    supplier.company_name = company_name
    supplier.company_address = (data.get("company_address") or "").strip() or None
    supplier.representative_name = (data.get("representative_name") or "").strip() or None
    supplier.contact_no = (data.get("contact_no") or "").strip() or None
    supplier.member_referrer_id = member_referrer_id
    supplier.date_joined = _parse_form_date(data.get("date_joined"), "date_joined") or supplier.date_joined


@main_routes.route("/admin/suppliers", methods=["POST"])
@login_required
@staff_or_admin_required
def admin_create_supplier():
    data = request.get_json() if request.is_json else request.form

    try:
        supplier_id = _parse_form_int(data.get("supplier_id"), "supplier_id")
        if supplier_id is None or supplier_id <= 0:
            raise ValueError("Supplier ID is required and must be greater than zero.")
        if db.session.get(Supplier, supplier_id):
            raise ValueError(f"Supplier #{supplier_id} already exists.")

        supplier = Supplier(supplier_id=supplier_id, batch=1, member_referrer_id=1, company_name="")
        _apply_supplier_form(supplier, data, include_batch=True)
        db.session.add(supplier)
        db.session.commit()
        return jsonify({
            "status": "success",
            "msg": f"Supplier #{supplier.supplier_id} added.",
            "supplier": supplier.to_dict(),
        })
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Database error: {exc}"}), 500


@main_routes.route("/admin/suppliers/<int:supplier_id>", methods=["POST"])
@login_required
@staff_or_admin_required
def admin_update_supplier(supplier_id):
    supplier = db.session.get(Supplier, supplier_id)
    if not supplier:
        return jsonify({"status": "error", "msg": "Supplier not found."}), 404

    data = request.get_json() if request.is_json else request.form

    try:
        _apply_supplier_form(supplier, data)
        db.session.commit()
        return jsonify({
            "status": "success",
            "msg": f"Supplier #{supplier.supplier_id} updated.",
            "supplier": supplier.to_dict(),
        })
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Database error: {exc}"}), 500


@main_routes.route("/admin/suppliers/<int:supplier_id>", methods=["DELETE"])
@login_required
@staff_or_admin_required
def admin_delete_supplier(supplier_id):
    supplier = db.session.get(Supplier, supplier_id)
    if not supplier:
        return jsonify({"status": "error", "msg": "Supplier not found."}), 404

    try:
        db.session.delete(supplier)
        db.session.commit()
        return jsonify({"status": "success", "msg": f"Supplier #{supplier_id} deleted."})
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Database error: {exc}"}), 500


@main_routes.route("/api/suppliers")
@login_required
def api_suppliers():
    denied = forbid_member_portal_users()
    if denied:
        return denied
    suppliers_list = (
        Supplier.query
        .options(joinedload(Supplier.member_referrer))
        .order_by(Supplier.company_name.asc())
        .all()
    )
    return jsonify([s.to_dict() for s in suppliers_list])


def _parse_form_date(value, field_name="date"):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Invalid {field_name}. Use YYYY-MM-DD.")


def _parse_form_int(value, field_name="value"):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(text)
    except ValueError as exc:
        raise ValueError(f"Invalid {field_name}. Must be a whole number.") from exc


def _parse_member_lifetime_cap(data):
    enabled = data.get("lifetime_cap_enabled") in ("1", "true", "True", "on", "yes")
    amount_raw = (data.get("lifetime_cap_amount") or "").strip().replace(",", "")
    if not amount_raw:
        amount = MEMBER_LIFETIME_EARNINGS_CAP
    else:
        amount = _parse_decimal(amount_raw, "lifetime_cap_amount")
    if amount < 0:
        raise ValueError("Lifetime cap amount cannot be negative.")
    return enabled, amount


def _apply_member_self_form(member, data):
    """Apply only fields members may edit on their own profile."""
    member.gender = (data.get("gender") or "").strip() or None
    member.civil_status = (data.get("civil_status") or "").strip() or None
    member.phone = (data.get("phone") or "").strip() or None
    member.email = (data.get("email") or "").strip() or None
    member.address = (data.get("address") or "").strip() or None
    member.highest_education = (data.get("highest_education") or "").strip() or None
    member.occupation_income_source = (data.get("occupation_income_source") or "").strip() or None
    member.monthly_income = (data.get("monthly_income") or "").strip() or None
    member.number_of_dependents = _parse_form_int(
        data.get("number_of_dependents"), "number_of_dependents"
    )
    member.beneficiary_name = (data.get("beneficiary_name") or "").strip() or None
    member.beneficiary_phone = (data.get("beneficiary_phone") or "").strip() or None
    member.beneficiary_address = (data.get("beneficiary_address") or "").strip() or None
    member.beneficiary_relationship = (data.get("beneficiary_relationship") or "").strip() or None
    member.gcash_number = (data.get("gcash_number") or "").strip() or None
    member.bank_account_number = (data.get("bank_account_number") or "").strip() or None
    member.bank_name = (data.get("bank_name") or "").strip() or None


def _apply_member_form(member, data, actor_role=None):
    first_name = (data.get("first_name") or "").strip()
    last_name = (data.get("last_name") or "").strip()
    if not first_name or not last_name:
        raise ValueError("First name and last name are required.")

    status = (data.get("status") or "Active").strip()
    if status not in MEMBER_STATUSES:
        raise ValueError(f"Invalid status. Allowed: {', '.join(MEMBER_STATUSES)}.")

    referrer_raw = (data.get("referrer_id") or "").strip()
    referrer_id = int(referrer_raw) if referrer_raw else None

    if referrer_id is None:
        if member.batch != 1:
            raise ValueError("Only batch 1 members can have no referrer.")
    else:
        if referrer_id == member.member_id:
            raise ValueError("Referrer cannot be the same member.")
        referrer = db.session.get(Member, referrer_id)
        if not referrer:
            raise ValueError("Referrer not found.")
        if referrer.batch >= member.batch:
            raise ValueError("Referrer must be from an earlier batch.")

    member.date_joined = _parse_form_date(data.get("date_joined"), "date_joined")
    member.first_name = first_name.upper()
    member.last_name = last_name.upper()
    member.middle_name = (data.get("middle_name") or "").strip() or None
    member.suffix = (data.get("suffix") or "").strip() or None
    member.address = (data.get("address") or "").strip() or None
    member.phone = (data.get("phone") or "").strip() or None
    member.email = (data.get("email") or "").strip() or None
    member.birth_date = _parse_form_date(data.get("birth_date"), "birth_date")
    member.gender = (data.get("gender") or "").strip() or None
    member.civil_status = (data.get("civil_status") or "").strip() or None
    member.highest_education = (data.get("highest_education") or "").strip() or None
    member.occupation_income_source = (data.get("occupation_income_source") or "").strip() or None
    member.monthly_income = (data.get("monthly_income") or "").strip() or None
    member.number_of_dependents = _parse_form_int(data.get("number_of_dependents"), "number_of_dependents")
    member.beneficiary_name = (data.get("beneficiary_name") or "").strip() or None
    member.beneficiary_address = (data.get("beneficiary_address") or "").strip() or None
    member.beneficiary_phone = (data.get("beneficiary_phone") or "").strip() or None
    member.beneficiary_relationship = (data.get("beneficiary_relationship") or "").strip() or None
    member.gcash_number = (data.get("gcash_number") or "").strip() or None
    member.bank_account_number = (data.get("bank_account_number") or "").strip() or None
    member.bank_name = (data.get("bank_name") or "").strip() or None
    member.age = _parse_form_int(data.get("age"), "age")
    member.id_picture_location = (data.get("id_picture_location") or "").strip() or None
    member.status = status
    member.termination_date = _parse_form_date(data.get("termination_date"), "termination_date")
    separation_type = (data.get("termination_type") or "").strip() or None
    if separation_type and separation_type not in MEMBER_SEPARATION_TYPES:
        raise ValueError(
            f"Invalid separation type. Allowed: {', '.join(MEMBER_SEPARATION_TYPES)}."
        )
    member.termination_type = separation_type
    member.referrer_id = referrer_id
    if is_admin_role(actor_role):
        member.lifetime_cap_enabled, member.lifetime_cap_amount = _parse_member_lifetime_cap(data)


@main_routes.route("/admin/members", methods=["POST"])
@login_required
@staff_or_admin_required
def admin_create_member():
    data = request.get_json() if request.is_json else request.form

    try:
        member_id = _parse_form_int(data.get("member_id"), "member_id")
        if member_id is None or member_id <= 0:
            raise ValueError("Member ID is required and must be greater than zero.")
        if db.session.get(Member, member_id):
            raise ValueError(f"Member #{member_id} already exists.")

        batch = _parse_form_int(data.get("batch"), "batch")
        if batch is None or batch <= 0:
            raise ValueError("Batch is required and must be greater than zero.")

        member = Member(member_id=member_id, batch=batch, first_name="", last_name="")
        _apply_member_form(member, data, actor_role=session.get("role"))
        db.session.add(member)
        db.session.commit()
        return jsonify({"status": "success", "msg": f"Member #{member.member_id} added.", "member": member.to_dict()})
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Database error: {exc}"}), 500


@main_routes.route("/admin/members/<int:member_id>", methods=["POST"])
@login_required
@staff_or_admin_required
def admin_update_member(member_id):
    member = db.session.get(Member, member_id)
    if not member:
        return jsonify({"status": "error", "msg": "Member not found."}), 404

    data = request.get_json() if request.is_json else request.form

    try:
        _apply_member_form(member, data, actor_role=session.get("role"))
        db.session.commit()
        return jsonify({"status": "success", "msg": f"Member #{member.member_id} updated.", "member": member.to_dict()})
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Database error: {exc}"}), 500


def _member_delete_blockers(member_id):
    checks = [
        ("direct referrals", Member.query.filter_by(referrer_id=member_id).count()),
        ("contractor referrals", Contractor.query.filter_by(member_referrer_id=member_id).count()),
        ("ledger transactions", MemberLedger.query.filter_by(member_id=member_id).count()),
        ("payout requests", PayoutRequest.query.filter_by(member_id=member_id).count()),
        ("OMPD fund entries", OmpdFundEntry.query.filter_by(member_id=member_id).count()),
        ("sharing entries", SharingEntry.query.filter_by(member_id=member_id).count()),
        ("linked user accounts", User.query.filter_by(member_id=member_id).count()),
        (
            "project commission references",
            ProjectCommission.query.filter(
                (ProjectCommission.client_referrer_id == member_id)
                | (ProjectCommission.contractor_referrer_id == member_id)
            ).count(),
        ),
    ]
    return [f"{count} {label}" for label, count in checks if count]


@main_routes.route("/admin/members/<int:member_id>", methods=["DELETE"])
@login_required
@staff_or_admin_required
def admin_delete_member(member_id):
    member = db.session.get(Member, member_id)
    if not member:
        return jsonify({"status": "error", "msg": "Member not found."}), 404

    blockers = _member_delete_blockers(member_id)
    if blockers:
        return jsonify({
            "status": "error",
            "msg": (
                "Cannot delete this member because they have related records: "
                f"{', '.join(blockers)}. Set the member status to Inactive or Separated instead."
            ),
        }), 400

    try:
        db.session.delete(member)
        db.session.commit()
        return jsonify({"status": "success", "msg": f"Member #{member_id} deleted."})
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Database error: {exc}"}), 500


@main_routes.route("/hierarchy")
@login_required
def hierarchy():
    user = User.query.get(session.get("user_id"))
    if is_member_role(user.role):
        linked_id = require_linked_member()
        if not linked_id:
            return access_denied_response("Your account is not linked to a member record.")
        tree = build_member_hierarchy_tree(linked_id)
        hierarchy_scope = "own"
    else:
        tree = build_hierarchy_tree()
        hierarchy_scope = "network"
    return render_template(
        "hierarchy.html",
        fullname=user.full_name or "Admin",
        role=normalize_role(user.role),
        tree=tree,
        hierarchy_scope=hierarchy_scope,
        active_page="hierarchy",
    )


@main_routes.route("/api/stats")
@login_required
def api_stats():
    if is_member_role(session.get("role")):
        linked_id = require_linked_member()
        if not linked_id:
            return jsonify({"error": "Not linked to a member record"}), 403
        return jsonify(member_dashboard_stats(linked_id))
    return jsonify(dashboard_stats())


@main_routes.route("/api/contractors")
@login_required
def api_contractors():
    denied = forbid_member_portal_users()
    if denied:
        return denied

    contractors_list = (
        Contractor.query
        .options(joinedload(Contractor.member_referrer))
        .order_by(Contractor.batch.asc(), Contractor.contractor_id.asc())
        .all()
    )
    return jsonify([c.to_dict() for c in contractors_list])


@main_routes.route("/api/members")
@login_required
def api_members():
    if is_member_role(session.get("role")):
        linked_id = require_linked_member()
        if not linked_id:
            return jsonify({"error": "Not linked to a member record"}), 403
        members_list = Member.query.filter_by(member_id=linked_id).all()
    else:
        members_list = Member.query.order_by(Member.batch.asc(), Member.member_id.asc()).all()
    return jsonify([m.to_dict() for m in members_list])


@main_routes.route("/api/hierarchy")
@login_required
def api_hierarchy():
    if is_member_role(session.get("role")):
        linked_id = require_linked_member()
        if not linked_id:
            return jsonify({"error": "Not linked to a member record"}), 403
        return jsonify(build_member_hierarchy_tree(linked_id))
    return jsonify(build_hierarchy_tree())


@main_routes.route("/api/member/<int:member_id>/lineage")
@login_required
def api_member_lineage(member_id):
    denied = forbid_unless_member_self(member_id)
    if denied:
        return denied
    data = member_lineage(member_id)
    if not data:
        return jsonify({"error": "Member not found"}), 404
    return jsonify(data)


@main_routes.route("/admin/reimport-contractors", methods=["POST"])
@login_required
@staff_or_admin_required
def reimport_contractors():
    try:
        result = import_contractors_from_xlsx(replace=True)
        return jsonify({"status": "success", "msg": "Contractors re-imported successfully.", **result})
    except Exception as exc:
        return jsonify({"status": "error", "msg": str(exc)}), 500


def _delete_all_members_and_dependents():
    """Delete member-dependent business data while preserving user accounts."""
    counts = {
        "ledger_entries": MemberLedger.query.count(),
        "payout_notifications": PayoutNotification.query.count(),
        "ompd_entries": OmpdFundEntry.query.count(),
        "payout_requests": PayoutRequest.query.count(),
        "sharing_entries": SharingEntry.query.count(),
        "sharing_batches": SharingBatch.query.count(),
        "project_billings": ProjectBilling.query.count(),
        "project_commissions": ProjectCommission.query.count(),
        "product_commissions": ProductCommission.query.count(),
        "ad_split_members": AdSplitMember.query.count(),
        "contractors": Contractor.query.count(),
        "suppliers": Supplier.query.count(),
        "members": Member.query.count(),
        "unlinked_users": User.query.filter(User.member_id.isnot(None)).count(),
    }

    # Shared with replace-import so purge cannot miss FK dependents (e.g. ad_split_members).
    clear_members_and_dependents(commit=False)
    for table_name, column_name in (
        ("members", "member_id"),
        ("contractors", "contractor_id"),
        ("suppliers", "supplier_id"),
    ):
        db.session.execute(text("""
            SELECT setval(sequence_name, 1, false)
            FROM pg_get_serial_sequence(:table_name, :column_name) AS sequence_name
            WHERE sequence_name IS NOT NULL
        """), {"table_name": table_name, "column_name": column_name})
    db.session.commit()
    return counts


@main_routes.route("/admin/purge-members", methods=["POST"])
@login_required
@admin_required
def purge_members():
    if not can_purge_member_database(session.get("role")):
        return jsonify({
            "status": "error",
            "msg": "Only the PortalAdmin account can purge member data.",
        }), 403

    data = request.get_json(silent=True) if request.is_json else request.form
    confirmation = (data.get("confirmation") if data else "") or ""
    if confirmation.strip() != "DELETE ALL MEMBERS":
        return jsonify({
            "status": "error",
            "msg": "Confirmation phrase did not match. Type DELETE ALL MEMBERS to continue.",
        }), 400

    try:
        counts = _delete_all_members_and_dependents()
        return jsonify({
            "status": "success",
            "msg": (
                f"Deleted {counts['members']} members, {counts['ledger_entries']} ledger entries, "
                f"{counts['contractors']} contractors, and related payout/sharing/project records. "
                "Member and contractor ID sequences were reset to 1."
            ),
            "counts": counts,
        })
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Purge failed: {exc}"}), 500


@main_routes.route("/admin/reimport", methods=["POST"])
@login_required
@staff_or_admin_required
def reimport_members():
    try:
        result = import_members_from_xlsx(replace=True, actor_role=session.get("role"))
        return jsonify({"status": "success", "msg": "Members re-imported successfully.", **result})
    except Exception as exc:
        return jsonify({"status": "error", "msg": str(exc)}), 500


@main_routes.route("/list_users")
@login_required
@staff_or_admin_required
def list_users():
    actor_role = session.get("role")
    users = (
        User.query
        .filter(User.username != "PortalAdmin")
        .filter(User.role != USER_ROLE_PORTAL_ADMIN)
        .order_by(User.user_id.asc())
        .all()
    )
    visible_users = [
        user for user in users
        if normalize_role(user.role) != USER_ROLE_PORTAL_ADMIN
    ]
    return jsonify({
        "users": [{
            "id": u.user_id,
            "username": u.username,
            "full_name": u.full_name,
            "role": u.role,
            "status": u.status,
            "member_id": u.member_id,
            "can_manage": staff_may_manage_user(actor_role, u.role),
        } for u in visible_users]
    })


@main_routes.route("/admin/add_user", methods=["POST"])
@login_required
@staff_or_admin_required
def add_user():
    username = request.form.get("username", "").strip()
    full_name = request.form.get("full_name", "").strip()
    role = normalize_role(request.form.get("role", "Staff"))
    password = request.form.get("password", "").strip()

    if not username or not password:
        return jsonify({"status": "error", "msg": "Username and password are required."})
    role_error = _role_assignment_error(session.get("role"), role)
    if role_error:
        return jsonify({"status": "error", "msg": role_error}), 403
    if User.query.filter_by(username=username).first():
        return jsonify({"status": "error", "msg": "Username already exists."})

    try:
        member_id = _parse_linked_member_id(role, request.form.get("member_id"))
    except ValueError as exc:
        return jsonify({"status": "error", "msg": str(exc)})

    user = User(
        username=username,
        full_name=full_name or username,
        role=role,
        status="Active",
        member_id=member_id,
        password_hash=generate_password_hash(password),
    )
    db.session.add(user)
    db.session.commit()
    return jsonify({"status": "success", "msg": f"User '{username}' added."})


@main_routes.route("/admin/update_user/<int:user_id>", methods=["POST"])
@login_required
@staff_or_admin_required
def update_user(user_id):
    user = User.query.get(user_id)
    if not user:
        return jsonify({"status": "error", "msg": "User not found"}), 404

    actor_role = session.get("role")
    if not staff_may_manage_user(actor_role, user.role):
        return jsonify({"status": "error", "msg": "You cannot modify this protected account."}), 403

    new_role = normalize_role(request.form.get("role", user.role))
    role_error = _role_assignment_error(actor_role, new_role)
    if role_error:
        return jsonify({"status": "error", "msg": role_error}), 403

    password = request.form.get("password", "").strip()
    if password:
        user.password_hash = generate_password_hash(password)

    user.username = request.form.get("username", user.username)
    user.full_name = request.form.get("full_name", user.full_name)
    user.role = new_role

    try:
        user.member_id = _parse_linked_member_id(user.role, request.form.get("member_id"))
    except ValueError as exc:
        return jsonify({"status": "error", "msg": str(exc)})

    db.session.commit()
    return jsonify({"status": "success", "msg": f"User '{user.username}' updated."})


@main_routes.route("/admin/delete_user", methods=["DELETE"])
@login_required
@staff_or_admin_required
def delete_user():
    data = request.get_json() or {}
    user = User.query.get(data.get("user_id"))
    if not user:
        return jsonify({"status": "error", "msg": "User not found"}), 404
    if user.username in ("Admin", "PortalAdmin"):
        return jsonify({"status": "error", "msg": "Cannot delete primary admin accounts."}), 400
    if not staff_may_manage_user(session.get("role"), user.role):
        return jsonify({"status": "error", "msg": "You cannot delete this protected account."}), 403

    db.session.delete(user)
    db.session.commit()
    return jsonify({"status": "success", "msg": f"User '{user.username}' deleted."})


def _dashboard_user():
    return User.query.get(session.get("user_id"))


def _parse_linked_member_id(role, raw_member_id):
    """Parse linked member ID. Required for Member; optional for Staff (My Marketplace)."""
    text = (raw_member_id or "").strip()
    if is_member_role(role):
        if not text.isdigit():
            raise ValueError("Member ID is required for Member role users.")
        member_id = int(text)
        if not db.session.get(Member, member_id):
            raise ValueError(f"Member #{member_id} not found.")
        return member_id
    if is_staff_role(role):
        if not text:
            return None
        if not text.isdigit():
            raise ValueError("Linked Member ID must be a number.")
        member_id = int(text)
        if not db.session.get(Member, member_id):
            raise ValueError(f"Member #{member_id} not found.")
        return member_id
    return None


def _role_assignment_error(actor_role, requested_role):
    role = normalize_role(requested_role)
    if role not in assignable_user_roles(actor_role):
        return "You cannot create or assign that role."
    return None


def _parse_decimal(value, field_name="amount"):
    if value is None:
        raise ValueError(f"{field_name} is required.")
    text = str(value).strip().replace(",", "")
    if not text:
        raise ValueError(f"{field_name} is required.")
    try:
        return round(float(text), 2)
    except ValueError as exc:
        raise ValueError(f"Invalid {field_name}.") from exc


def _parse_billings_payload(data):
    raw = data.get("billings")
    if raw is None:
        raw = data.get("billings_json")
    if isinstance(raw, str):
        text = raw.strip()
        raw = json.loads(text) if text else []
    if not isinstance(raw, list) or not raw:
        raise ValueError("At least one billing date and amount is required.")

    billings = []
    for index, row in enumerate(raw, start=1):
        if not isinstance(row, dict):
            raise ValueError(f"Invalid billing row #{index}.")
        billing_date = _parse_form_date(row.get("billing_date"), f"billing #{index} date")
        if not billing_date:
            raise ValueError(f"Billing #{index} date is required.")
        billing_amount = _parse_decimal(row.get("billing_amount"), f"billing #{index} amount")
        billing_id = row.get("billing_id")
        billings.append({
            "billing_id": int(billing_id) if billing_id not in (None, "") else None,
            "billing_date": billing_date,
            "billing_amount": billing_amount,
        })
    return billings


def _same_money(left, right):
    return round(float(left or 0), 2) == round(float(right or 0), 2)


def _billing_generated_sharing_count(billing_id):
    return SharingEntry.query.filter_by(billing_id=billing_id).count()


def _sync_project_billings(project, billings_data, protect_generated_amounts=False):
    existing = {billing.billing_id: billing for billing in list(project.billings)}
    keep_ids = set()
    for row in billings_data:
        billing_id = row["billing_id"]
        if billing_id and billing_id in existing:
            billing = existing[billing_id]
            sharing_count = _billing_generated_sharing_count(billing_id)
            if protect_generated_amounts and sharing_count and not _same_money(
                billing.billing_amount,
                row["billing_amount"],
            ):
                raise ValueError(
                    "Staff cannot edit the amount of a billing that already has generated sharing. "
                    "Ask an Admin to change it, or delete the related generated sharing batch first."
                )
            billing.billing_date = row["billing_date"]
            billing.billing_amount = row["billing_amount"]
            keep_ids.add(billing_id)
        else:
            db.session.add(ProjectBilling(
                project_id=project.project_id,
                billing_date=row["billing_date"],
                billing_amount=row["billing_amount"],
            ))
    for billing_id, billing in existing.items():
        if billing_id not in keep_ids:
            sharing_count = SharingEntry.query.filter_by(billing_id=billing_id).count()
            if sharing_count:
                raise ValueError(
                    "Cannot delete this billing because generated sharing already references it. "
                    "Delete the related generated sharing batch first, then edit the billing rows."
                )
            db.session.delete(billing)


def _project_generated_sharing_count(project_id):
    return SharingEntry.query.filter_by(project_id=project_id).count()


def _staff_generated_project_locked(project, actor_role):
    return bool(
        project
        and is_staff_role(actor_role)
        and _project_generated_sharing_count(project.project_id)
    )


def _project_generated_sharing_counts(project_ids):
    if not project_ids:
        return {}
    rows = (
        db.session.query(SharingEntry.project_id, func.count(SharingEntry.entry_id))
        .filter(SharingEntry.project_id.in_(project_ids))
        .group_by(SharingEntry.project_id)
        .all()
    )
    return {project_id: int(count) for project_id, count in rows}


def _billing_generated_sharing_counts(billing_ids):
    if not billing_ids:
        return {}
    rows = (
        db.session.query(SharingEntry.billing_id, func.count(SharingEntry.entry_id))
        .filter(SharingEntry.billing_id.in_(billing_ids))
        .group_by(SharingEntry.billing_id)
        .all()
    )
    return {billing_id: int(count) for billing_id, count in rows}


def _project_commission_payloads(projects):
    project_ids = [project.project_id for project in projects]
    billing_ids = [
        billing.billing_id
        for project in projects
        for billing in project.billings
        if billing.billing_id is not None
    ]
    project_counts = _project_generated_sharing_counts(project_ids)
    billing_counts = _billing_generated_sharing_counts(billing_ids)

    payloads = {}
    for project in projects:
        payload = project.to_dict()
        payload["generated_sharing_count"] = project_counts.get(project.project_id, 0)
        for billing in payload["billings"]:
            billing["generated_sharing_count"] = billing_counts.get(billing["billing_id"], 0)
        payloads[project.project_id] = payload
    return payloads


@main_routes.route("/admin/prof/project-commission")
@login_required
@staff_or_admin_required
def prof_project_commission():
    user = _dashboard_user()
    projects = (
        ProjectCommission.query
        .options(
            joinedload(ProjectCommission.contractor),
            joinedload(ProjectCommission.client_referrer),
            joinedload(ProjectCommission.contractor_referrer),
            joinedload(ProjectCommission.billings),
        )
        .order_by(ProjectCommission.project_id.desc())
        .all()
    )
    contractors = Contractor.query.order_by(Contractor.company_name.asc()).all()
    members = Member.query.order_by(Member.last_name.asc(), Member.first_name.asc()).all()
    contractor_referrer_map = {
        str(c.contractor_id): c.member_referrer_id for c in contractors
    }
    member_label_map = {
        str(m.member_id): ProjectCommission.member_referral_label(m) for m in members
    }
    return render_template(
        "prof_project_commission.html",
        fullname=user.full_name or "Admin",
        role=normalize_role(user.role),
        active_page="prof_project_commission",
        projects=projects,
        project_payloads=_project_commission_payloads(projects),
        contractors=contractors,
        members=members,
        contractor_referrer_map=contractor_referrer_map,
        member_label_map=member_label_map,
    )


@main_routes.route("/admin/prof/project-commission", methods=["POST"])
@login_required
@staff_or_admin_required
def prof_project_commission_save():
    data = request.get_json() if request.is_json else request.form
    project_id = (data.get("project_id") or "").strip()
    actor_role = session.get("role")

    try:
        project = None
        if project_id:
            project = db.session.get(ProjectCommission, int(project_id))
            if not project:
                return jsonify({"status": "error", "msg": "Project not found."}), 404

        staff_generated_lock = _staff_generated_project_locked(project, actor_role)

        if staff_generated_lock:
            submitted_title = (data.get("project_title") or "").strip()
            submitted_address = (data.get("address") or "").strip() or None
            submitted_contractor_id = (data.get("contractor_id") or "").strip()
            submitted_client_referrer_id = (data.get("client_referrer_id") or "").strip()

            if submitted_title and submitted_title != project.project_title:
                raise ValueError("Staff cannot edit the project title after sharing has been generated.")
            if submitted_address != project.address:
                raise ValueError("Staff cannot edit the project address after sharing has been generated.")
            if submitted_contractor_id and int(submitted_contractor_id) != project.contractor_id:
                raise ValueError("Staff cannot edit the contractor after sharing has been generated.")
            if submitted_client_referrer_id and int(submitted_client_referrer_id) != project.client_referrer_id:
                raise ValueError("Staff cannot edit the project client referral after sharing has been generated.")

            contractor_id = project.contractor_id
            client_referrer_id = project.client_referrer_id
            project_title = project.project_title
            project_address = project.address
        else:
            contractor_id = int(data.get("contractor_id"))
            client_referrer_raw = (data.get("client_referrer_id") or "").strip()
            if not client_referrer_raw:
                raise ValueError("Project client referral is required.")
            client_referrer_id = int(client_referrer_raw)
            project_title = (data.get("project_title") or "").strip()
            project_address = (data.get("address") or "").strip() or None

        billings_data = _parse_billings_payload(data)

        contractor = db.session.get(Contractor, contractor_id)
        if not contractor:
            raise ValueError("Contractor not found.")

        client_referrer = db.session.get(Member, client_referrer_id)
        if not client_referrer:
            raise ValueError("Project client referral member not found.")

        contractor_referrer_id = (
            project.contractor_referrer_id if staff_generated_lock else contractor.member_referrer_id
        )
        if not contractor_referrer_id:
            raise ValueError("Selected contractor has no member referrer.")
        if not db.session.get(Member, contractor_referrer_id):
            raise ValueError("Contractor member referrer not found.")

        payload = {
            "project_title": project_title,
            "address": project_address,
            "contractor_id": contractor_id,
            "client_referrer_id": client_referrer_id,
            "contractor_referrer_id": contractor_referrer_id,
        }
        if not payload["project_title"]:
            raise ValueError("Project title is required.")

        if project_id:
            for key, value in payload.items():
                setattr(project, key, value)
            msg = "Project commission updated."
        else:
            project = ProjectCommission(**payload)
            db.session.add(project)
            db.session.flush()
            msg = "Project commission added."

        _sync_project_billings(
            project,
            billings_data,
            protect_generated_amounts=staff_generated_lock,
        )
        db.session.commit()
        return jsonify({"status": "success", "msg": msg, "project": project.to_dict()})
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Save failed: {exc}"}), 500


@main_routes.route("/admin/prof/project-commission/<int:project_id>", methods=["DELETE"])
@login_required
@staff_or_admin_required
def prof_project_commission_delete(project_id):
    project = db.session.get(ProjectCommission, project_id)
    if not project:
        return jsonify({"status": "error", "msg": "Project not found."}), 404
    sharing_count = _project_generated_sharing_count(project_id)
    if sharing_count:
        return jsonify({
            "status": "error",
            "msg": (
                "Cannot delete this project commission because generated sharing already references "
                f"{sharing_count} row(s). Delete the related generated sharing batch first."
            ),
        }), 400
    try:
        db.session.delete(project)
        db.session.commit()
        return jsonify({"status": "success", "msg": "Project commission deleted."})
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Delete failed: {exc}"}), 500


@main_routes.route("/admin/prof/products-commission")
@login_required
@admin_required
def prof_products_commission():
    user = _dashboard_user()
    products = (
        ProductCommission.query
        .options(
            joinedload(ProductCommission.ref_seller),
            joinedload(ProductCommission.ref_buyer),
            joinedload(ProductCommission.shares).joinedload(ProductCommissionShare.member),
            joinedload(ProductCommission.ad_allocations).joinedload(
                ProductCommissionAdAllocation.member
            ),
        )
        .order_by(
            ProductCommission.commission_date.desc(),
            ProductCommission.product_commission_id.desc(),
        )
        .all()
    )
    members_list = (
        Member.query
        .filter(Member.status == "Active")
        .order_by(Member.batch.asc(), Member.member_id.asc())
        .all()
    )
    ad_split_members = (
        AdSplitMember.query
        .options(joinedload(AdSplitMember.member))
        .order_by(AdSplitMember.ad_split_id.asc())
        .all()
    )
    return render_template(
        "prof_products_commission.html",
        fullname=user.full_name or "Admin",
        role=normalize_role(user.role),
        active_page="prof_products_commission",
        products=products,
        members=members_list,
        ad_split_members=ad_split_members,
        product_bonus_auto_percent=float(PRODUCT_BONUS_AUTO_PERCENT),
        product_bonus_ad_default_percent=float(PRODUCT_BONUS_AD_DEFAULT_PERCENT),
        product_split={
            "ref_seller": float(PRODUCT_REF_SELLER_PERCENT),
            "ref_buyer": float(PRODUCT_REF_BUYER_PERCENT),
            "pop_pct": float(PRODUCT_POP_PERCENT),
            "ad_fund": float(PRODUCT_AD_FUND_PERCENT),
            "platform": float(PRODUCT_PLATFORM_PERCENT),
        },
        product_pool_levels=[
            {"level": level, "percentage": float(pct), "description": desc}
            for level, pct, desc in DEFAULT_PRODUCT_POOL_LEVELS
        ],
    )


@main_routes.route("/admin/prof/products-commission/preview", methods=["POST"])
@login_required
@admin_required
def prof_products_commission_preview():
    data = request.get_json() if request.is_json else request.form
    try:
        commission_date = _parse_form_date(data.get("commission_date"), "commission_date")
        if not commission_date:
            raise ValueError("Commission date is required.")
        amount = _parse_decimal(data.get("commission_amount"), "commission_amount")
        ref_seller_id = _parse_form_int(data.get("ref_seller_id"), "ref_seller_id")
        ref_buyer_id = _parse_form_int(data.get("ref_buyer_id"), "ref_buyer_id")
        if not ref_seller_id or not ref_buyer_id:
            raise ValueError("Ref-Seller and Ref-Buyer are required.")
        result = compute_product_commission(
            commission_amount=amount,
            ref_seller_id=ref_seller_id,
            ref_buyer_id=ref_buyer_id,
            bonus_type=data.get("bonus_type") or PRODUCT_BONUS_TYPE_AUTO,
            ad_bonus_percent=data.get("ad_bonus_percent"),
            product_title=(data.get("product_title") or "").strip() or "Products Commission",
            commission_date=commission_date,
            ad_allocations=data.get("ad_allocations"),
        )
        result.pop("_share_objects", None)
        result.pop("_allocation_objects", None)
        return jsonify({"status": "success", "computation": result})
    except ValueError as exc:
        return jsonify({"status": "error", "msg": str(exc)}), 400
    except Exception as exc:
        return jsonify({"status": "error", "msg": f"Preview failed: {exc}"}), 500


@main_routes.route("/admin/prof/products-commission", methods=["POST"])
@login_required
@admin_required
def prof_products_commission_save():
    data = request.get_json() if request.is_json else request.form
    try:
        commission_date = _parse_form_date(data.get("commission_date"), "commission_date")
        if not commission_date:
            raise ValueError("Commission date is required.")
        amount = _parse_decimal(data.get("commission_amount"), "commission_amount")
        ref_seller_id = _parse_form_int(data.get("ref_seller_id"), "ref_seller_id")
        ref_buyer_id = _parse_form_int(data.get("ref_buyer_id"), "ref_buyer_id")
        if not ref_seller_id or not ref_buyer_id:
            raise ValueError("Ref-Seller and Ref-Buyer are required.")

        product = save_product_commission(
            commission_amount=amount,
            ref_seller_id=ref_seller_id,
            ref_buyer_id=ref_buyer_id,
            bonus_type=data.get("bonus_type") or PRODUCT_BONUS_TYPE_AUTO,
            ad_bonus_percent=data.get("ad_bonus_percent"),
            product_title=(data.get("product_title") or "").strip() or "Products Commission",
            commission_date=commission_date,
            notes=data.get("notes"),
            created_by_user_id=session.get("user_id"),
            ad_allocations=data.get("ad_allocations"),
        )
        return jsonify({
            "status": "success",
            "msg": f"Products commission #{product.product_commission_id} saved.",
            "product": product.to_dict(),
        })
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Save failed: {exc}"}), 500


@main_routes.route("/admin/prof/products-commission/<int:product_commission_id>", methods=["DELETE"])
@login_required
@admin_required
def prof_products_commission_delete(product_commission_id):
    try:
        delete_product_commission(product_commission_id)
        return jsonify({"status": "success", "msg": "Products commission deleted."})
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Delete failed: {exc}"}), 500


@main_routes.route("/admin/prof/commission-levels")
@login_required
@admin_required
def prof_commission_levels():
    user = _dashboard_user()
    client_levels = get_commission_levels(COMMISSION_SCHEME_CLIENT)
    contractor_levels = get_commission_levels(COMMISSION_SCHEME_CONTRACTOR)
    client_total_pct = sum(float(row.percentage or 0) for row in client_levels)
    contractor_total_pct = sum(float(row.percentage or 0) for row in contractor_levels)
    ad_split_members = (
        AdSplitMember.query
        .options(joinedload(AdSplitMember.member))
        .order_by(AdSplitMember.ad_split_id.asc())
        .all()
    )
    active_members = (
        Member.query
        .filter(Member.status == "Active")
        .order_by(Member.batch.asc(), Member.member_id.asc())
        .all()
    )
    return render_template(
        "prof_commission_levels.html",
        fullname=user.full_name or "Admin",
        role=normalize_role(user.role),
        active_page="prof_commission_levels",
        client_levels=client_levels,
        contractor_levels=contractor_levels,
        client_total_pct=client_total_pct,
        contractor_total_pct=contractor_total_pct,
        ad_split_members=ad_split_members,
        members=active_members,
    )


@main_routes.route("/admin/prof/ad-split-members", methods=["POST"])
@login_required
@admin_required
def prof_ad_split_members_save():
    data = request.get_json() if request.is_json else request.form
    ad_split_id = (data.get("ad_split_id") or "").strip()

    try:
        member_id = _parse_form_int(data.get("member_id"), "member_id")
        if not member_id:
            raise ValueError("Member is required.")
        member = db.session.get(Member, member_id)
        if not member:
            raise ValueError("Member not found.")
        if member.status != "Active":
            raise ValueError("Only active members can be added to AD-Members Split Sharing.")

        description = (data.get("description") or "").strip() or None
        if description and len(description) > 255:
            raise ValueError("Description must be 255 characters or fewer.")

        if ad_split_id:
            row = db.session.get(AdSplitMember, int(ad_split_id))
            if not row:
                return jsonify({"status": "error", "msg": "AD-Member entry not found."}), 404
            conflict = AdSplitMember.query.filter(
                AdSplitMember.member_id == member_id,
                AdSplitMember.ad_split_id != row.ad_split_id,
            ).first()
            if conflict:
                raise ValueError(f"Member #{member_id} is already in AD-Members Split Sharing.")
            row.member_id = member_id
            row.description = description
            row.updated_at = datetime.utcnow()
            msg = "AD-Member updated."
        else:
            if AdSplitMember.query.filter_by(member_id=member_id).first():
                raise ValueError(f"Member #{member_id} is already in AD-Members Split Sharing.")
            row = AdSplitMember(
                member_id=member_id,
                description=description,
                created_at=datetime.utcnow(),
                updated_at=datetime.utcnow(),
                created_by_user_id=session.get("user_id"),
            )
            db.session.add(row)
            msg = "AD-Member added."

        db.session.commit()
        return jsonify({"status": "success", "msg": msg, "ad_member": row.to_dict()})
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Save failed: {exc}"}), 500


@main_routes.route("/admin/prof/ad-split-members/<int:ad_split_id>", methods=["DELETE"])
@login_required
@admin_required
def prof_ad_split_members_delete(ad_split_id):
    row = db.session.get(AdSplitMember, ad_split_id)
    if not row:
        return jsonify({"status": "error", "msg": "AD-Member entry not found."}), 404
    try:
        db.session.delete(row)
        db.session.commit()
        return jsonify({"status": "success", "msg": "AD-Member removed."})
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Delete failed: {exc}"}), 500


@main_routes.route("/admin/prof/commission-levels", methods=["POST"])
@login_required
@admin_required
def prof_commission_levels_save():
    data = request.get_json() if request.is_json else request.form
    level_id = (data.get("level_id") or "").strip()

    try:
        scheme = (data.get("scheme") or COMMISSION_SCHEME_CLIENT).strip()
        if scheme not in COMMISSION_SCHEMES:
            raise ValueError("Invalid commission scheme.")

        level_no = int(data.get("level"))
        percentage = _parse_decimal(data.get("percentage"), "percentage")
        if level_no < 1 or level_no > MAX_SHARING_LEVELS:
            raise ValueError(f"Level must be between 1 and {MAX_SHARING_LEVELS}.")
        if percentage < 0 or percentage > 100:
            raise ValueError("Percentage must be between 0 and 100.")

        description = (data.get("description") or "").strip() or None
        if level_no == MAX_SHARING_LEVELS:
            description = description or "Mandate account (Level 7)"
        elif level_no == 1 and not description:
            description = (
                "Project client referrer"
                if scheme == COMMISSION_SCHEME_CLIENT
                else "Contractor member referrer"
            )

        if level_id:
            row = db.session.get(CommissionLevel, int(level_id))
            if not row:
                return jsonify({"status": "error", "msg": "Level not found."}), 404
            conflict = CommissionLevel.query.filter(
                CommissionLevel.scheme == scheme,
                CommissionLevel.level == level_no,
                CommissionLevel.level_id != row.level_id,
            ).first()
            if conflict:
                raise ValueError(f"Level {level_no} already exists for this scheme.")
            row.scheme = scheme
            row.level = level_no
            row.percentage = percentage
            row.description = description
            msg = "Commission level updated."
        else:
            if CommissionLevel.query.filter_by(scheme=scheme, level=level_no).first():
                raise ValueError(f"Level {level_no} already exists for this scheme.")
            row = CommissionLevel(
                scheme=scheme,
                level=level_no,
                percentage=percentage,
                description=description,
            )
            db.session.add(row)
            msg = "Commission level added."

        db.session.commit()
        return jsonify({"status": "success", "msg": msg, "level": row.to_dict()})
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Save failed: {exc}"}), 500


@main_routes.route("/admin/prof/commission-levels/<int:level_id>", methods=["DELETE"])
@login_required
@admin_required
def prof_commission_levels_delete(level_id):
    row = db.session.get(CommissionLevel, level_id)
    if not row:
        return jsonify({"status": "error", "msg": "Level not found."}), 404
    if row.level == MAX_SHARING_LEVELS:
        return jsonify({"status": "error", "msg": "Level 7 (Mandate pool) cannot be deleted."}), 400
    try:
        db.session.delete(row)
        db.session.commit()
        return jsonify({"status": "success", "msg": "Commission level deleted."})
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Delete failed: {exc}"}), 500


@main_routes.route("/members/ledger")
@login_required
def member_ledger():
    user = _dashboard_user()
    member_id_raw = request.args.get("member_id", "").strip()
    member_id = int(member_id_raw) if member_id_raw.isdigit() else None

    if is_member_role(user.role):
        linked_id = require_linked_member()
        if not linked_id:
            return access_denied_response("Your account is not linked to a member record.")
        if member_id and member_id != linked_id:
            return access_denied_response("You may only view your own ledger.")
        member_id = linked_id
        members_list = []
    else:
        members_list = Member.query.order_by(Member.last_name.asc(), Member.first_name.asc()).all()

    stats = member_ledger_stats(member_id)
    transactions = member_ledger_rows(member_id)
    selected_member = db.session.get(Member, member_id) if member_id else None
    payout_requests = []
    if member_id and is_member_role(user.role):
        payout_requests = [
            payout_to_dict(row)
            for row in payout_request_query(member_id=member_id).limit(10).all()
        ]
    return render_template(
        "member_ledger.html",
        fullname=user.full_name or "User",
        role=normalize_role(user.role),
        active_page="member_ledger",
        members=members_list,
        selected_member=selected_member,
        selected_member_id=member_id,
        stats=stats,
        transactions=transactions,
        is_own_ledger_view=is_member_role(user.role),
        payout_requests=payout_requests,
        payout_release_methods=PAYOUT_RELEASE_METHODS,
        can_request_payout=can_request_payout(user.role) and bool(member_id),
        payout_scheme=payout_scheme_summary(),
    )


@main_routes.route("/api/members/<int:member_id>/ledger")
@login_required
def api_member_ledger(member_id):
    denied = forbid_unless_own_ledger(member_id)
    if denied:
        return denied
    member = db.session.get(Member, member_id)
    if not member:
        return jsonify({"status": "error", "msg": "Member not found."}), 404
    limit = request.args.get("limit", type=int) or 10
    return jsonify({
        "status": "success",
        "member_id": member_id,
        "member_name": member.full_name,
        "stats": member_ledger_stats(member_id),
        "transactions": member_ledger_rows(member_id, limit=limit),
    })


@main_routes.route("/admin/prof/generate-sharing")
@login_required
@staff_or_admin_required
def prof_generate_sharing():
    user = _dashboard_user()
    pending_dates = ungenerated_billing_dates()
    generated_dates = generated_billing_dates()
    client_levels = get_commission_levels(COMMISSION_SCHEME_CLIENT)
    contractor_levels = get_commission_levels(COMMISSION_SCHEME_CONTRACTOR)
    selected_date = request.args.get("commission_date", "")
    preview_billings = []
    latest_batch = None

    if selected_date:
        try:
            billing_date = _parse_form_date(selected_date, "commission_date")
            preview_billings = billings_for_billing_date(billing_date)
            if is_admin_role(user.role):
                latest_batch = (
                    SharingBatch.query
                    .filter_by(commission_date=billing_date)
                    .order_by(SharingBatch.generated_at.desc())
                    .first()
                )
        except ValueError:
            selected_date = ""

    batch_detail = sharing_batch_summary(latest_batch) if latest_batch else None
    history_batches = (
        SharingBatch.query.order_by(SharingBatch.generated_at.desc()).limit(20).all()
        if is_admin_role(user.role)
        else []
    )
    can_generate_selected = (
        selected_date
        and selected_date in {d.isoformat() for d in pending_dates}
    )

    return render_template(
        "prof_generate_sharing.html",
        fullname=user.full_name or "Admin",
        role=normalize_role(user.role),
        active_page="prof_generate_sharing",
        pending_dates=pending_dates,
        generated_dates=generated_dates,
        selected_date=selected_date,
        can_generate_selected=can_generate_selected,
        preview_billings=preview_billings,
        client_levels=client_levels,
        contractor_levels=contractor_levels,
        batches=history_batches,
        batch_detail=batch_detail,
    )


@main_routes.route("/admin/prof/generate-sharing", methods=["POST"])
@login_required
@staff_or_admin_required
def prof_generate_sharing_run():
    data = request.get_json() if request.is_json else request.form
    try:
        billing_date = _parse_form_date(data.get("commission_date"), "commission_date")
        if not billing_date:
            raise ValueError("Billing date is required.")
        batch = generate_profit_sharing(billing_date)
        payload = {
            "status": "success",
            "msg": f"Profit sharing generated for {billing_date.isoformat()}.",
        }
        if is_admin_role(session.get("role")):
            payload["batch"] = sharing_batch_summary(batch)
        return jsonify(payload)
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Generation failed: {exc}"}), 500


@main_routes.route("/admin/prof/generate-sharing/batch/<int:batch_id>", methods=["DELETE"])
@login_required
@admin_required
def prof_generate_sharing_delete(batch_id):
    try:
        billing_date = remove_sharing_batch(batch_id)
        return jsonify({
            "status": "success",
            "msg": f"Deleted sharing for {billing_date.isoformat()}. Billing date is available to generate again.",
        })
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "msg": f"Delete failed: {exc}"}), 500


@main_routes.route("/admin/prof/reports/project-list")
@login_required
@admin_required
def prof_reports_project_list():
    user = _dashboard_user()
    projects = generated_projects_index()
    selected_id = request.args.get("project_id", type=int)
    report = project_detail_report(selected_id) if selected_id else None
    return render_template(
        "prof_reports_project_list.html",
        fullname=user.full_name or "Admin",
        role=normalize_role(user.role),
        active_page="prof_reports_project_list",
        projects=projects,
        selected_project_id=selected_id,
        report=report,
    )


@main_routes.route("/admin/prof/reports/project-list/<int:project_id>/pdf")
@login_required
@admin_required
def prof_reports_project_pdf(project_id):
    report = project_detail_report(project_id)
    if not report:
        flash("Project report not found or sharing has not been generated.", "warning")
        return redirect(url_for("main_routes.prof_reports_project_list"))

    pdf_buffer = build_project_report_pdf(report)
    filename = f"project-report-{project_id}.pdf"
    return send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )


@main_routes.route("/admin/prof/reports/commission-summary")
@login_required
@admin_required
def prof_reports_commission_summary():
    user = _dashboard_user()
    report = commission_summary_report()
    return render_template(
        "prof_reports_commission_summary.html",
        fullname=user.full_name or "Admin",
        role=normalize_role(user.role),
        active_page="prof_reports_commission_summary",
        report=report,
    )


@main_routes.route("/admin/prof/reports/commission-summary/pdf")
@login_required
@admin_required
def prof_reports_commission_summary_pdf():
    report = commission_summary_report()
    if not report.get("has_data"):
        flash("No commission summary data available yet.", "warning")
        return redirect(url_for("main_routes.prof_reports_commission_summary"))

    pdf_buffer = build_commission_summary_pdf(report)
    return send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name="commission-summary.pdf",
    )


def _parse_report_datetime(value, end_of_day=False):
    if not value:
        return None
    text = value.strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            parsed = datetime.strptime(text, fmt)
            if end_of_day and fmt == "%Y-%m-%d":
                return parsed.replace(hour=23, minute=59, second=59)
            return parsed
        except ValueError:
            continue
    raise ValueError(f"Invalid date/time: {value}")


@main_routes.route("/payout/scheme")
@login_required
@staff_or_admin_required
def payout_scheme():
    user = _dashboard_user()
    if not can_view_payout_scheme(user.role):
        return access_denied_response("You cannot view the payout scheme.")

    summary = ompd_fund_summary()
    recent_entries = ompd_fund_rows(limit=25)
    return render_template(
        "payout_scheme.html",
        fullname=user.full_name or "User",
        role=normalize_role(user.role),
        active_page="payout_scheme",
        scheme=payout_scheme_summary(),
        fund_summary=summary,
        recent_entries=recent_entries,
    )


@main_routes.route("/members/ledger/payout-request", methods=["POST"])
@login_required
def member_payout_request():
    user = _dashboard_user()
    if not can_request_payout(user.role):
        return access_denied_response("You cannot request payouts.")

    linked_id = require_linked_member()
    if not linked_id:
        return access_denied_response("Your account is not linked to a member record.")

    amount_raw = request.form.get("amount", "").strip().replace(",", "")
    note = request.form.get("member_note", "").strip()
    try:
        amount = float(amount_raw)
        create_payout_request(linked_id, amount, user.user_id, note)
        flash("Payout request submitted. Staff and Admin have been notified.", "success")
    except ValueError as exc:
        flash(str(exc), "danger")
    except Exception as exc:
        db.session.rollback()
        flash(f"Request failed: {exc}", "danger")

    return redirect(url_for("main_routes.member_ledger"))


@main_routes.route("/payout/queue")
@login_required
@staff_or_admin_required
def payout_queue():
    user = _dashboard_user()
    tab = request.args.get("tab", "pending").strip().lower()
    status_map = {
        "pending": PAYOUT_STATUS_PENDING,
        "approved": PAYOUT_STATUS_APPROVED,
        "release": PAYOUT_STATUS_RELEASE_SUBMITTED,
        "released": PAYOUT_STATUS_RELEASED,
    }
    status = status_map.get(tab)
    rows = payout_request_query(status=status).all() if status else payout_request_query().all()
    return render_template(
        "payout_queue.html",
        fullname=user.full_name or "User",
        role=normalize_role(user.role),
        active_page="payout_queue",
        tab=tab,
        payouts=[payout_to_dict(row) for row in rows],
        payout_release_methods=PAYOUT_RELEASE_METHODS,
        can_approve_payout_request=can_approve_payout_request(user.role),
        can_submit_payout_release=can_submit_payout_release(user.role),
        can_approve_payout_release=can_approve_payout_release(user.role),
        payout_scheme=payout_scheme_summary(),
    )


@main_routes.route("/payout/requests/<int:payout_id>/approve", methods=["POST"])
@login_required
@admin_required
def payout_request_approve(payout_id):
    user = _dashboard_user()
    note = request.form.get("note", "").strip()
    try:
        approve_payout_request(payout_id, user.user_id, note)
        flash("Payout request approved.", "success")
    except ValueError as exc:
        flash(str(exc), "danger")
    except Exception as exc:
        db.session.rollback()
        flash(f"Approval failed: {exc}", "danger")
    return redirect(url_for("main_routes.payout_queue", tab="pending"))


@main_routes.route("/payout/requests/<int:payout_id>/reject", methods=["POST"])
@login_required
@admin_required
def payout_request_reject(payout_id):
    user = _dashboard_user()
    reason = request.form.get("reason", "").strip()
    try:
        reject_payout_request(payout_id, user.user_id, reason)
        flash("Payout request rejected.", "warning")
    except ValueError as exc:
        flash(str(exc), "danger")
    except Exception as exc:
        db.session.rollback()
        flash(f"Rejection failed: {exc}", "danger")
    return redirect(url_for("main_routes.payout_queue", tab="pending"))


@main_routes.route("/payout/requests/<int:payout_id>/release", methods=["POST"])
@login_required
@staff_or_admin_required
def payout_request_release(payout_id):
    user = _dashboard_user()
    if not can_submit_payout_release(user.role):
        return access_denied_response("You cannot submit fund releases.")
    try:
        submit_payout_release(
            payout_id=payout_id,
            staff_user_id=user.user_id,
            method=request.form.get("release_method", ""),
            reference=request.form.get("release_reference", ""),
            account_info=request.form.get("release_account_info", ""),
            notes=request.form.get("release_notes", ""),
            other_method=request.form.get("release_method_other", ""),
            bank_name=request.form.get("release_bank_name", ""),
            bank_branch=request.form.get("release_bank_branch", ""),
        )
        flash("Fund release submitted for admin approval.", "success")
    except ValueError as exc:
        flash(str(exc), "danger")
    except Exception as exc:
        db.session.rollback()
        flash(f"Release submission failed: {exc}", "danger")
    return redirect(url_for("main_routes.payout_queue", tab="approved"))


@main_routes.route("/payout/requests/<int:payout_id>/approve-release", methods=["POST"])
@login_required
@admin_required
def payout_release_approve(payout_id):
    user = _dashboard_user()
    try:
        approve_payout_release(payout_id, user.user_id)
        flash("Fund release approved and recorded on member ledger.", "success")
    except ValueError as exc:
        flash(str(exc), "danger")
    except Exception as exc:
        db.session.rollback()
        flash(f"Release approval failed: {exc}", "danger")
    return redirect(url_for("main_routes.payout_queue", tab="release"))


@main_routes.route("/payout/requests/<int:payout_id>/reject-release", methods=["POST"])
@login_required
@admin_required
def payout_release_reject(payout_id):
    user = _dashboard_user()
    reason = request.form.get("reason", "").strip()
    try:
        reject_payout_release(payout_id, user.user_id, reason)
        flash("Fund release rejected. Staff may resubmit.", "warning")
    except ValueError as exc:
        flash(str(exc), "danger")
    except Exception as exc:
        db.session.rollback()
        flash(f"Release rejection failed: {exc}", "danger")
    return redirect(url_for("main_routes.payout_queue", tab="release"))


@main_routes.route("/payout/reports")
@login_required
@staff_or_admin_required
def payout_reports():
    user = _dashboard_user()
    if not can_view_payout_reports(user.role):
        return access_denied_response("You cannot view payout reports.")

    member_id_raw = request.args.get("member_id", "").strip()
    member_id = int(member_id_raw) if member_id_raw.isdigit() else None
    method = request.args.get("method", "").strip() or None
    date_from_raw = request.args.get("date_from", "").strip()
    date_to_raw = request.args.get("date_to", "").strip()

    try:
        date_from = _parse_report_datetime(date_from_raw) if date_from_raw else None
        date_to = _parse_report_datetime(date_to_raw, end_of_day=True) if date_to_raw else None
    except ValueError as exc:
        flash(str(exc), "danger")
        date_from = None
        date_to = None

    rows = fund_release_rows(date_from, date_to, member_id, method)
    summary = fund_release_summary(date_from, date_to, member_id, method)
    members_list = Member.query.order_by(Member.last_name.asc(), Member.first_name.asc()).all()

    return render_template(
        "payout_reports.html",
        fullname=user.full_name or "User",
        role=normalize_role(user.role),
        active_page="payout_reports",
        rows=rows,
        summary=summary,
        members=members_list,
        filters={
            "member_id": member_id,
            "method": method or "",
            "date_from": date_from_raw,
            "date_to": date_to_raw,
        },
        payout_release_methods=PAYOUT_RELEASE_METHODS,
        payout_scheme=payout_scheme_summary(),
    )


@main_routes.route("/payout/reports/pdf")
@login_required
@staff_or_admin_required
def payout_reports_pdf():
    if not can_view_payout_reports(session.get("role")):
        return access_denied_response("You cannot view payout reports.")

    member_id_raw = request.args.get("member_id", "").strip()
    member_id = int(member_id_raw) if member_id_raw.isdigit() else None
    method = request.args.get("method", "").strip() or None
    date_from_raw = request.args.get("date_from", "").strip()
    date_to_raw = request.args.get("date_to", "").strip()

    try:
        date_from = _parse_report_datetime(date_from_raw) if date_from_raw else None
        date_to = _parse_report_datetime(date_to_raw, end_of_day=True) if date_to_raw else None
    except ValueError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("main_routes.payout_reports"))

    rows = fund_release_rows(date_from, date_to, member_id, method)
    if not rows:
        flash("No fund release records match your filters.", "warning")
        return redirect(url_for("main_routes.payout_reports"))

    summary = fund_release_summary(date_from, date_to, member_id, method)
    pdf_buffer = build_fund_release_pdf(rows, summary, {
        "date_from": date_from_raw,
        "date_to": date_to_raw,
        "method": method or "",
        "member_id": member_id,
    })
    return send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name="fund-release-report.pdf",
    )
